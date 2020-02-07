from openpyxl import Workbook
import pandas
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import xlrd
import sys
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THIN, BORDER_THICK
from openpyxl.styles import Font


wb=Workbook()
ws1= wb.create_sheet("Sheet_1")
ws9= wb.create_sheet("Sheet_9")
ws2= wb.create_sheet("Sheet_2")
ws4= wb.create_sheet("Sheet_4")
ws5= wb.create_sheet("Sheet_5")
ws6= wb.create_sheet("Sheet_6")
ws7= wb.create_sheet("Sheet_7")

#name the tabs
ws1.title="Patient demographics"
ws2.title="Variant_calls"
ws4.title="Mutations and SNPs"
ws5.title="hotspots.gaps"
ws6.title="Report"
ws7.title="NTC variant"
ws9.title="Subpanel NTC check"


#set the page layout of the report
ws6.page_setup.orientation=ws6.ORIENTATION_LANDSCAPE
ws6.page_setup.paperSize=ws6.PAPERSIZE_A4


#Patient demographics tab table headers
ws1['A1']='Date Received'
ws1['B1']='Leeds/Cardiff'
ws1['C1']='Lab No'
ws1['D1']='Notes'
ws1['E1']='Patient Name'
ws1['F1']='Tumour %'
ws1['G1']='Qubit [DNA] ng/u'
ws1['H1']='Dilution (ng/ul)'
ws1['I1']='NGS wks'
ws1['J1']='Date Set up'
ws1['K1']='Date of MiSeq run'
ws1['L1']='Library ng/ul'
ws1['M1']='Library nM'

#variant calls table headers
ws2['B3']='DNA number'
ws2['B6']='Patient name'

ws2['E3']='NTC check 1'
ws2['E6']='NTC check 2'

ws2['G3']='1st checker name & date'
ws2['G6']='2nd checker name & date'

ws2['K3']='CRM worksheet'
ws2['K6']='Analysis pipeline:CRM'
ws2['A8']=" "

ws2['I3']= "% Tumour"


#Mutations and SNPs table headers
ws4['B2']='Gene'
ws4['C2']='Exon/Intron'
ws4['D2']='HGVS c'
ws4['E2']='HGVS p'
ws4['F2']='Allele Frequency'
ws4['G2']='Quality'
ws4['H2']='Depth'
ws4['I2']='Classification'
ws4['J2']='Transcript'
ws4['K2']='Variant'
ws4['L2']='Position'
ws4['M2']='Conclusion 1st checker'
ws4['N2']='Conclusion 2nd checker'


#Report headers
ws6['A3']='Patient information'
ws6['A4']='Lab number'
ws6['B4']='Patient Name'
ws6['C4']='Tumour %'
ws6['D4']='Analysis'
ws6['E4']='Qubit[DNA] ng/ul'
ws6['F4']='Dilution (ng/ul)'
ws6['G4']='Analysed by:'
ws6['H4']='Checked by:'


ws6['E7']='NGS wks'
ws6['F7']='Date set up'
ws6['G7']='Date of MiSeq run'
ws6['H7']='Library ng/ul(Qubit)'
ws6['I7']='Library nm'
ws6['J7']='NTC check 1'
ws6['K7']='NTC check 2'

ws6['A4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['B4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['C4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['D4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['E4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['F4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['G4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['H4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['E7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['F7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['G7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['H7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['I7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['J7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['K7'].fill= PatternFill("solid", fgColor="00CCFFFF")


ws6['A11']='Confirmed variant calls'
ws6['A12']='Gene'
ws6['B12']='Exon'
ws6['C12']='Variant'
ws6['D12']='HGVS c.'
ws6['E12']='HGVS p.'
ws6['F12']='Allele frequency'
ws6['G12']='Conclusion 1st checker'
ws6['H12']='Conclusion 2nd checker'

ws6['A12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['B12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['C12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['D12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['E12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['F12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['G12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['H12'].fill= PatternFill("solid", fgColor="009999FF")

    


def get_variantReport_NTC(referral, path): 
    '''
    Fill out the NTC variants tab using the relevant variant report
    '''

    if(os.stat(path+NTC_name+"/hotspot_variants/"+runid+"_"+NTC_name+"_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report_NTC=pandas.read_csv(path+NTC_name+"/hotspot_variants/"+runid+"_"+NTC_name+"_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(path+NTC_name+"/hotspot_variants/"+runid+"_"+NTC_name+"_"+referral+"_VariantReport.txt")
    else:
        variant_report_NTC=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "PreferredTranscript","dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
 

    variant_report_NTC_2=pandas.DataFrame(variant_report_NTC)

    #Sort by preferred transcript and filter columns
    variant_report_NTC_3=variant_report_NTC_2[variant_report_NTC_2.Preferred!=False]
    variant_report_NTC_4= variant_report_NTC_3.iloc[:,[23,29,25,26,2,5,3,6,24,1]]

    return (variant_report_NTC_4)




def get_variant_report(referral, path, sampleid):

    '''
    Open the relevant variant file to append to the variant calls tab.
    '''

    if(os.stat(path+ sampleid +"/hotspot_variants/"+runid+"_"+sampleid+"_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report=pandas.read_csv(path+sampleid+"/hotspot_variants/" +runid+"_"+sampleid+"_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(sampleid+"_"+referral+"_VariantReport.txt")
    else:
        variant_report=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "PreferredTranscript", "dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
            

    #sort dataframe by preferred transcript an filter out certain columns
    variant_report_2=pandas.DataFrame(variant_report)

    variant_report_3=variant_report_2[variant_report_2.Preferred!=False]
    variant_report_4= variant_report_3.iloc[:,[23,29,25,26,2,5,3,6,24,1]]
    
    #Add position column in variant calls tab by splitting the variant column 
    variant_list=[]    
    for variant in variant_report_4['Variant']:
        variant_2=[]
        value=0
        for char in variant:
            if(value==0):
                if (char.isdigit()==True or char=="X" or char==":"):
                    variant_2.append(char)
                else:
                    value=1
        variant_3=''.join(variant_2)
        variant_list.append(variant_3)
        

    variant_report_4['Position']=variant_list


    return (variant_report_4)





def add_extra_columns_NTC_report(variant_report_NTC_4, variant_report_4):
    
    '''
    Add 'Present in sample' and 'variant allele calls' columns to NTC variant table
    '''



    num_rows_NTC=variant_report_NTC_4.shape[0]
    num_rows_variant_report=variant_report_4.shape[0]

    variant_in_sample=[]
    row=0
    while (row<num_rows_NTC):
        row2=0
        present='NO'
        while (row2<num_rows_variant_report):
            if (variant_report_4.iloc[row2,9]==variant_report_NTC_4.iloc[row,9]):
                present='YES'
            row2=row2+1
        variant_in_sample.append(present)
        row=row+1
     
   
    variant_report_NTC_4['Present in sample']=variant_in_sample

    variant_allele_calls=[]
    row=0
    num_rows_NTC=variant_report_NTC_4.shape[0]
    while (row<num_rows_NTC):
        variant_report_NTC_4.iloc[row,6]=int(variant_report_NTC_4.iloc[row,6])
        allele_call= variant_report_NTC_4.iloc[row,4]*variant_report_NTC_4.iloc[row,6]
        variant_allele_calls.append(allele_call)
        row=row+1

    variant_report_NTC_4['Variant allele calls']=variant_allele_calls    

    for row in dataframe_to_rows(variant_report_NTC_4, header=True, index=False):
        ws7.append(row)

    return (variant_report_NTC_4)




def expand_variant_report(variant_report_4, variant_report_NTC_4):

    '''
    create the extra table at the side of the variant_calls tab
    '''

    detection_threshold=[]
    variant_in_NTC=[]


    row=0
    num_rows_variant_report=variant_report_4.shape[0]


    while (row<num_rows_variant_report):
        variant_report_4.iloc[row,6]= int(variant_report_4.iloc[row,6])

        if (variant_report_4.iloc[row,6]<=500):
            value_2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            value_2=str(value_2)
        else:
            value_2=0
        detection_threshold.append(value_2)
        row2=0
        num_rows_NTC=variant_report_NTC_4.shape[0]
        present='NO'
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row,9]==variant_report_NTC_4.iloc[row2,9]):
                present='YES'
            row2=row2+1
        variant_in_NTC.append(present)
        row=row+1


    variant_report_4["Conclusion 1st checker"]=""
    variant_report_4["QC"]=""
    variant_report_4["Conclusion 2nd checker"]=""
    variant_report_4["QC "]=""
    variant_report_4[""]=""
    
    if (referral=="FOCUS4"):
        variant_report_4["Variant classification if seen before"]=""
    else:
        variant_report_4[""]=""

    variant_report_4["Detection threshold based on depth"]=detection_threshold
    variant_report_4["Is variant present in NTC "]=variant_in_NTC


    return (variant_report_4)




def get_gaps_file(referral, path, sampleid):
    
    ''' 
    Open the relevant gap file to append to the end of the mutations and snps tab. If the gap file is empty, write 'no gaps'.
    '''
    hotspot_variants=referral
    #open the relevant bed files (ready for screening gaps tab)
    if (hotspot_variants=='GIST'):
        if((os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KIT.gaps")==0) and (os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PDGFRA.gaps")==0)):
            ws5['A1']= 'No gaps'
        if (os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KIT.gaps")>0):
            bedfile_KIT=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KIT.gaps", sep="\t")
            for row in dataframe_to_rows(bedfile_KIT, header=True, index=False):
                ws5.append(row)
        if (os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PDGFRA.gaps")>0):
            bedfile_PDGFRA=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PDGFRA.gaps", sep="\t")
            for row in dataframe_to_rows(bedfile_PDGFRA, header=True, index=False):
                ws5.append(row)
    elif (hotspot_variants=='FOCUS4'):
        if((os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_BRAF.gaps")==0) and (os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KRAS.gaps")==0) and (os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_NRAS.gaps")==0) and (os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PIK3CA.gaps")==0)and (os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_TP53.gaps")==0)):
            ws5['A1']= 'No gaps'
        if((os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_BRAF.gaps")>0)):
            bedfile_BRAF=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_BRAF.gaps", sep="\t")
            for row in dataframe_to_rows(bedfile_BRAF, header=True, index=False):
                ws5.append(row)
        if((os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KRAS.gaps")>0)):
            bedfile_KRAS=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KRAS.gaps", sep="\t")
            for row in dataframe_to_rows(bedfile_KRAS, header=True, index=False):
                ws5.append(row)
        if((os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_NRAS.gaps")>0)):
            bedfile_NRAS=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_NRAS.gaps", sep="\t")
            for row in dataframe_to_rows(bedfile_NRAS, header=True, index=False):
                ws5.append(row)
        if((os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PIK3CA.gaps")>0)):
            bedfile_PIK3CA=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PIK3CA.gaps", sep="\t")
            for row in dataframe_to_rows(bedfile_PIK3CA, header=True, index=False):
                ws5.append(row)
        if((os.path.getsize(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_TP53.gaps")>0)):
            bedfile_TP53=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_TP53.gaps", sep="\t")
            for row in dataframe_to_rows(bedfile_TP53, header=True, index=False):
                ws5.append(row)



    ws6['E30']="=hotspots.gaps!D1"
    ws6['E31']="=hotspots.gaps!D2"
    ws6['E32']="=hotspots.gaps!D3"
    ws6['E33']="=hotspots.gaps!D4"
    ws6['E34']="=hotspots.gaps!D5"
    ws6['E35']="=hotspots.gaps!D6"
    ws6['E36']="=hotspots.gaps!D7"
    ws6['E37']="=hotspots.gaps!D8"
    ws6['E38']="=hotspots.gaps!D9"
    ws6['E39']="=hotspots.gaps!D10"
    ws6['E40']="=hotspots.gaps!D11"
    ws6['E41']="=hotspots.gaps!D12"
    ws6['E42']="=hotspots.gaps!D13"

    ws6['E43']="=hotspots.gaps!D14"
    ws6['E44']="=hotspots.gaps!D15"
    ws6['F40']="=hotspots.gaps!D16"
    ws6['F30']="=hotspots.gaps!D17"
    ws6['F31']="=hotspots.gaps!D18"
    ws6['F32']="=hotspots.gaps!D19"
    ws6['F33']="=hotspots.gaps!D20"
    ws6['F34']="=hotspots.gaps!D21"
    ws6['F35']="=hotspots.gaps!D22"
    ws6['F36']="=hotspots.gaps!D23"
    ws6['F37']="=hotspots.gaps!D24"
    ws6['F38']="=hotspots.gaps!D25"
    ws6['F39']="=hotspots.gaps!D26"
    ws6['F40']="=hotspots.gaps!D27"
    ws6['F41']="=hotspots.gaps!D28"
    ws6['F42']="=hotspots.gaps!D29"
    ws6['F43']="=hotspots.gaps!D30"
    ws6['F44']="=hotspots.gaps!D31"





def get_hotspots_coverage_file(referral, path, sampleid):
    '''
    Open the relevant coverage file to append to the end of the mutations and snps tab. If the coverage file is empty, write 'No hotspots'.
    '''
    hotspot_variants=referral
    if (hotspot_variants=='GIST'):
        bedfile_KIT=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KIT.coverage", sep="\t")
        bedfile_PDGFRA=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PDGFRA.coverage", sep="\t")
        Coverage=pandas.concat([bedfile_KIT, bedfile_PDGFRA])
    elif (hotspot_variants=='FOCUS4'):
        bedfile_BRAF=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_BRAF.coverage", sep="\t")
        bedfile_KRAS=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_KRAS.coverage", sep="\t")
        bedfile_NRAS=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_NRAS.coverage", sep="\t")
        bedfile_PIK3CA=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_PIK3CA.coverage", sep="\t")
        bedfile_TP53=pandas.read_csv(path+ sampleid+ "/hotspot_coverage/"+runid+"_"+sampleid+"_TP53.coverage", sep="\t")
        Coverage=pandas.concat([bedfile_BRAF, bedfile_KRAS, bedfile_NRAS, bedfile_PIK3CA, bedfile_TP53])


    Coverage= Coverage.iloc[:,[3,4,5]]
    


    return(Coverage)




def get_NTC_hotspots_coverage_file(referral, path):

    '''
    Open the relevant NTC hotspots coverage file.
    '''
    hotspot_variants=referral       
    if (hotspot_variants=='GIST'):
        bedfile_NTC_KIT=pandas.read_csv(path+NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_KIT.coverage", sep="\t")
        bedfile_NTC_PDGFRA=pandas.read_csv(path+NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_PDGFRA.coverage", sep="\t")
        NTC_check= pandas.concat([bedfile_NTC_KIT, bedfile_NTC_PDGFRA]) 
    elif (hotspot_variants=='FOCUS4'):
        bedfile_NTC_BRAF=pandas.read_csv(path+ NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_BRAF.coverage", sep="\t")
        bedfile_NTC_KRAS=pandas.read_csv(path+ NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_KRAS.coverage", sep="\t")
        bedfile_NTC_NRAS=pandas.read_csv(path+ NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_NRAS.coverage", sep="\t")
        bedfile_NTC_PIK3CA=pandas.read_csv(path+ NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_PIK3CA.coverage", sep="\t")
        bedfile_NTC_TP53=pandas.read_csv(path+NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_TP53.coverage", sep="\t")
        NTC_check=pandas.concat([bedfile_NTC_BRAF, bedfile_NTC_KRAS, bedfile_NTC_NRAS, bedfile_NTC_PIK3CA, bedfile_NTC_TP53])
    return(NTC_check)




def add_columns_hotspots_coverage(Coverage, NTC_check):

    #Add percentage NTC and subpanel columns to the Coverage table
    
    Coverage['NTC_AVG_Depth']=""
    Coverage['%NTC']=""

    num_rows_NTC= NTC_check.shape[0]
    num_rows_sample= Coverage.shape[0]

    row1=0

    while(row1<num_rows_sample):
        row2=0
        while (row2<num_rows_NTC):
            if(Coverage.iloc[row1,0] == NTC_check.iloc[row2,3]):
                Coverage.iloc[row1,3]= NTC_check.iloc[row2,4]
            row2=row2+1
        row1=row1+1
    Coverage['%NTC']=Coverage['NTC_AVG_Depth']/Coverage['AVG_DEPTH']
    Coverage['%NTC']= Coverage['%NTC']*100

    for row in dataframe_to_rows(Coverage, header=True, index=False):
        ws9.append(row)

    num_rows_coverage=Coverage.shape[0]
    row =0
    while (row< num_rows_coverage):
        row_spreadsheet=row+2
        row_spreadsheet_2=str(row_spreadsheet)
        if (Coverage.iloc[row,4]>10):
            ws9['E'+row_spreadsheet_2].fill= PatternFill("solid", fgColor="FFBB00")
        row=row+1
    
    return (Coverage, num_rows_coverage)




def match_polys_and_artefacts(variant_report_4, variant_report_NTC_4):

    '''
    Extract the relevant information from "PanCancer_Poly and Artefact list.xlsx" by matching the variant name with the ones in the variant report table
    '''

    poly_artefact_dict={}
    poly_and_Artefact_list_2=pandas.read_excel("/data/temp/artefacts_lists/CRM_poly_artefact_list.xlsx")
    variant_spreadsheet=pandas.read_excel("/data/temp/artefacts_lists/FOCUS_4_Variants.xlsx",sheet_name="Variants")
    num_rows_variant_report=variant_report_4.shape[0]
    num_rows_poly_artefact=poly_and_Artefact_list_2.shape[0]



    #Fill the conclusion columns using the relevant column in the Poly and Artefact spreadsheet
    row1=0
    while (row1<num_rows_variant_report):
        row2=0
        while(row2<num_rows_poly_artefact):
            if (poly_and_Artefact_list_2.iloc[row2,0]==variant_report_4.iloc[row1,9]):
                variant_report_4.iloc[row1,11]= poly_and_Artefact_list_2.iloc[row2,9]
                variant_report_4.iloc[row1,13]= poly_and_Artefact_list_2.iloc[row2,9]
            row2=row2+1
        row1=row1+1

    print(variant_report_4)

    #fill second table of variant-calls tab using the conclusion column of the first table
    row3=0
    while (row3<num_rows_variant_report):
        if (variant_report_4.iloc[row3,11]=='Known Artefact'):
            variant_report_4.iloc[row3,12]=3
            variant_report_4.iloc[row3,14]=3
        if (variant_report_4.iloc[row3,11]=='Known Poly'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc[row3,14]=1
        if (variant_report_4.iloc[row3,11]=='WT'):
            variant_report_4.iloc[row3,12]=3
            variant_report_4.iloc[row3,14]=3
        if (variant_report_4.iloc[row3,11]=='Genuine'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc_[row3,14]=1
        if (variant_report_4.iloc[row3,11]=='SNP'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc_[row3,14]=1

        row3=row3+1

    
    #Match variants to the variants list to determine what its classification was before
    if (referral=="FOCUS4"):
        num_rows_variant_spreadsheet=variant_spreadsheet.shape[0]
        row1=0
        while (row1<num_rows_variant_report):
            row2=0
            while(row2<num_rows_variant_spreadsheet):
                if (variant_spreadsheet.iloc[row2,11]==variant_report_4.iloc[row1,9]):
                    variant_report_4.iloc[row1,16]= variant_spreadsheet.iloc[row2,12]
                row2=row2+1
            row1=row1+1



 

   #Add extra columns to the variant report table to determine level of NTC contamination

    
    variant_report_4["#of mutant reads in patient sample "]=""
    variant_report_4["#of mutant reads in NTC if present "]=""
    variant_report_4["Is the NTC contamination significant?"]=""


    num_rows_NTC= variant_report_NTC_4.shape[0]
    row=0


    while (row<num_rows_variant_report):
        if variant_report_4.iloc[row,16]=="YES":
            variant_report_4.iloc[row,6]=float(variant_report_4.iloc[row,6])
            variant_report_4.iloc[row,4]=float(variant_report_4.iloc[row,4])
            value2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            variant_report_4.iloc[row,17]=value2
        row2=0
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row, 9]==variant_report_NTC_4.iloc[row2,9]):
                variant_report_4.iloc[row,18]=variant_report_NTC_4.iloc[row2,11]
                variant_report_4.iloc[row,19]=variant_report_4.iloc[row,18]/variant_report_4.iloc[row,17]
            row2=row2+1
        row=row+1


   
    for row in dataframe_to_rows(variant_report_4, header=True, index=False):
        ws2.append(row)


    variant_report_5= variant_report_4.iloc[:,[0,1,2]]
    variant_report_5['Comments/Notes/evidence:how conclusion was reached']=""



    row=0

    num_rows_variant_report=variant_report_4.shape[0]

    while (row<num_rows_variant_report):
        if (variant_report_4.iloc[row,11]=='Known artefact'):
            variant_report_5.iloc[row,3]='On artefact list'
        if (variant_report_4.iloc[row,11]=='Known Poly'):
            variant_report_5.iloc[row,3]='On Poly list'
        if (variant_report_4.iloc[row,11]=='WT'):
            variant_report_5.iloc[row,3]='SNP in Ref.Seq'
        row=row+1

    ws2['A60']=" "




    #add dataframe to variant calls tab

    for row in dataframe_to_rows(variant_report_5,  header=True, index=False):
        ws2.append(row)


    return(variant_report_4)



def add_excel_formulae():

    #add excel formulae to the spreadsheets to enable automation after program has finished

    ws2['I4']= "='Patient demographics'!F2"
     
    ws6['A13']= "='Mutations and SNPS'!B3"
    ws6['B13']= "='Mutations and SNPS'!C3"
    ws6['C13']= "='Mutations and SNPS'!K3"
    ws6['D13']= "='Mutations and SNPS'!D3"
    ws6['E13']= "='Mutations and SNPS'!E3"
    ws6['F13']= "='Mutations and SNPS'!F3"
    ws6['G13']= "='Mutations and SNPS'!M3"
    ws6['H13']= "='Mutations and SNPS'!N3"

    ws6['A14']= "='Mutations and SNPS'!B4"
    ws6['B14']= "='Mutations and SNPS'!C4"
    ws6['C14']= "='Mutations and SNPS'!K4"
    ws6['D14']= "='Mutations and SNPS'!D4"
    ws6['E14']= "='Mutations and SNPS'!E4"
    ws6['F14']= "='Mutations and SNPS'!F4"
    ws6['G14']= "='Mutations and SNPS'!M4"
    ws6['H14']= "='Mutations and SNPS'!N4"

    ws6['A15']= "='Mutations and SNPS'!B5"
    ws6['B15']= "='Mutations and SNPS'!C5"
    ws6['C15']= "='Mutations and SNPS'!K5"
    ws6['D15']= "='Mutations and SNPS'!D5"
    ws6['E15']= "='Mutations and SNPS'!E5"
    ws6['F15']= "='Mutations and SNPS'!F5"
    ws6['G15']= "='Mutations and SNPS'!M5"
    ws6['H15']= "='Mutations and SNPS'!N5"

    ws6['A16']= "='Mutations and SNPS'!B6"
    ws6['B16']= "='Mutations and SNPS'!C6"
    ws6['C16']= "='Mutations and SNPS'!K6"
    ws6['D16']= "='Mutations and SNPS'!D6"
    ws6['E16']= "='Mutations and SNPS'!E6"
    ws6['F16']= "='Mutations and SNPS'!F6"
    ws6['G16']= "='Mutations and SNPS'!M6"
    ws6['H16']= "='Mutations and SNPS'!N6"

    ws6['A17']= "='Mutations and SNPS'!B7"
    ws6['B17']= "='Mutations and SNPS'!C7"
    ws6['C17']= "='Mutations and SNPS'!K7"
    ws6['D17']= "='Mutations and SNPS'!D7"
    ws6['E17']= "='Mutations and SNPS'!E7"
    ws6['F17']= "='Mutations and SNPS'!F7"
    ws6['G17']= "='Mutations and SNPS'!M7"
    ws6['H17']= "='Mutations and SNPS'!N7"

    ws6['A18']= "='Mutations and SNPS'!B8"
    ws6['B18']= "='Mutations and SNPS'!C8"
    ws6['C18']= "='Mutations and SNPS'!K8"
    ws6['D18']= "='Mutations and SNPS'!D8"
    ws6['E18']= "='Mutations and SNPS'!E8"
    ws6['F18']= "='Mutations and SNPS'!F8"
    ws6['G18']= "='Mutations and SNPS'!M8"
    ws6['H18']= "='Mutations and SNPS'!N8"

    ws6['A19']= "='Mutations and SNPS'!B9"
    ws6['B19']= "='Mutations and SNPS'!C9"
    ws6['C19']= "='Mutations and SNPS'!K9"
    ws6['D19']= "='Mutations and SNPS'!D9"
    ws6['E19']= "='Mutations and SNPS'!E9"
    ws6['F19']= "='Mutations and SNPS'!F9"
    ws6['G19']= "='Mutations and SNPS'!M9"
    ws6['H19']= "='Mutations and SNPS'!N9"

    ws6['A20']= "='Mutations and SNPS'!B10"
    ws6['B20']= "='Mutations and SNPS'!C10"
    ws6['C20']= "='Mutations and SNPS'!K10"
    ws6['D20']= "='Mutations and SNPS'!D10"
    ws6['E20']= "='Mutations and SNPS'!E10"
    ws6['F20']= "='Mutations and SNPS'!F10"
    ws6['G20']= "='Mutations and SNPS'!M10"
    ws6['H20']= "='Mutations and SNPS'!N10"

    ws6['A21']= "='Mutations and SNPS'!B11"
    ws6['B21']= "='Mutations and SNPS'!C11"
    ws6['C21']= "='Mutations and SNPS'!K11"
    ws6['D21']= "='Mutations and SNPS'!D11"
    ws6['E21']= "='Mutations and SNPS'!E11"
    ws6['F21']= "='Mutations and SNPS'!F11"
    ws6['G21']= "='Mutations and SNPS'!M11"
    ws6['H21']= "='Mutations and SNPS'!N11"

    ws6['A22']= "='Mutations and SNPS'!B12"
    ws6['B22']= "='Mutations and SNPS'!C12"
    ws6['C22']= "='Mutations and SNPS'!K12"
    ws6['D22']= "='Mutations and SNPS'!D12"
    ws6['E22']= "='Mutations and SNPS'!E12"
    ws6['F22']= "='Mutations and SNPS'!F12"
    ws6['G22']= "='Mutations and SNPS'!M12"
    ws6['H22']= "='Mutations and SNPS'!N12"

    ws6['A23']= "='Mutations and SNPS'!B13"
    ws6['B23']= "='Mutations and SNPS'!C13"
    ws6['C23']= "='Mutations and SNPS'!K13"
    ws6['D23']= "='Mutations and SNPS'!D13"
    ws6['E23']= "='Mutations and SNPS'!E13"
    ws6['F23']= "='Mutations and SNPS'!F13"
    ws6['G23']= "='Mutations and SNPS'!M13"
    ws6['H23']= "='Mutations and SNPS'!N13"

    ws6['A24']= "='Mutations and SNPS'!B14"
    ws6['B24']= "='Mutations and SNPS'!C14"
    ws6['C24']= "='Mutations and SNPS'!K14"
    ws6['D24']= "='Mutations and SNPS'!D14"
    ws6['E24']= "='Mutations and SNPS'!E14"
    ws6['F24']= "='Mutations and SNPS'!F14"
    ws6['G24']= "='Mutations and SNPS'!M14"
    ws6['H24']= "='Mutations and SNPS'!N14"

    ws6['A25']= "='Mutations and SNPS'!B15"
    ws6['B25']= "='Mutations and SNPS'!C15"
    ws6['C25']= "='Mutations and SNPS'!K15"
    ws6['D25']= "='Mutations and SNPS'!D15"
    ws6['E25']= "='Mutations and SNPS'!E15"
    ws6['F25']= "='Mutations and SNPS'!F15"
    ws6['G25']= "='Mutations and SNPS'!M15"
    ws6['H25']= "='Mutations and SNPS'!N15"

    ws6['A26']= "='Mutations and SNPS'!B16"
    ws6['B26']= "='Mutations and SNPS'!C16"
    ws6['C26']= "='Mutations and SNPS'!K16"
    ws6['D26']= "='Mutations and SNPS'!D16"
    ws6['E26']= "='Mutations and SNPS'!E16"
    ws6['F26']= "='Mutations and SNPS'!F16"
    ws6['G26']= "='Mutations and SNPS'!M16"
    ws6['H26']= "='Mutations and SNPS'!N16"


    ws6['A29']= "Regions"
    ws6['B29'] = "% bases"
    ws6['C29']= "Regions"               
    ws6['D29'] = "% bases"
    ws6['E29']= "Hotspots.gaps"
    ws6['G29']= "Comments"

    ws6['A30']= "='SUBPANEL NTC CHECK'!A2"
    ws6['A31']= "='SUBPANEL NTC CHECK'!A3"
    ws6['A32']= "='SUBPANEL NTC CHECK'!A4"
    ws6['A33']= "='SUBPANEL NTC CHECK'!A5"
    ws6['A34']= "='SUBPANEL NTC CHECK'!A6"
    ws6['A35']= "='SUBPANEL NTC CHECK'!A7"
    ws6['A36']= "='SUBPANEL NTC CHECK'!A8"
    ws6['A37']= "='SUBPANEL NTC CHECK'!A9"
    ws6['A38']= "='SUBPANEL NTC CHECK'!A10"
    ws6['A39']= "='SUBPANEL NTC CHECK'!A11"
    ws6['A40']= "='SUBPANEL NTC CHECK'!A12"
    ws6['A41']= "='SUBPANEL NTC CHECK'!A13"
    ws6['A42']= "='SUBPANEL NTC CHECK'!A14"
    ws6['A43']= "='SUBPANEL NTC CHECK'!A15"
    ws6['A44']= "='SUBPANEL NTC CHECK'!A16"

    ws6['C30']= "='SUBPANEL NTC CHECK'!A17"
    ws6['C31']= "='SUBPANEL NTC CHECK'!A18"
    ws6['C32']= "='SUBPANEL NTC CHECK'!A19"
    ws6['C33']= "='SUBPANEL NTC CHECK'!A20"
    ws6['C34']= "='SUBPANEL NTC CHECK'!A21"
    ws6['C35']= "='SUBPANEL NTC CHECK'!A22"
    ws6['C36']= "='SUBPANEL NTC CHECK'!A23"
    ws6['C37']= "='SUBPANEL NTC CHECK'!A24"
    ws6['C38']= "='SUBPANEL NTC CHECK'!A25"
    ws6['C39']= "='SUBPANEL NTC CHECK'!A26"
    ws6['C40']= "='SUBPANEL NTC CHECK'!A27"
    ws6['C41']= "='SUBPANEL NTC CHECK'!A28"
    ws6['C42']= "='SUBPANEL NTC CHECK'!A29"
    ws6['C43']= "='SUBPANEL NTC CHECK'!A30"
    ws6['C44']= "='SUBPANEL NTC CHECK'!A31"



    ws6['B30']= "='SUBPANEL NTC CHECK'!C2"
    ws6['B31']= "='SUBPANEL NTC CHECK'!C3"
    ws6['B32']= "='SUBPANEL NTC CHECK'!C4"
    ws6['B33']= "='SUBPANEL NTC CHECK'!C5"
    ws6['B34']= "='SUBPANEL NTC CHECK'!C6"
    ws6['B35']= "='SUBPANEL NTC CHECK'!C7"
    ws6['B36']= "='SUBPANEL NTC CHECK'!C8"
    ws6['B37']= "='SUBPANEL NTC CHECK'!C9"
    ws6['B38']= "='SUBPANEL NTC CHECK'!C10"
    ws6['B39']= "='SUBPANEL NTC CHECK'!C11"
    ws6['B40']= "='SUBPANEL NTC CHECK'!C12"
    ws6['B41']= "='SUBPANEL NTC CHECK'!C13"
    ws6['B42']= "='SUBPANEL NTC CHECK'!C14"
    ws6['B43']= "='SUBPANEL NTC CHECK'!C15"
    ws6['B44']= "='SUBPANEL NTC CHECK'!C16"

    ws6['D30']= "='SUBPANEL NTC CHECK'!C17"
    ws6['D31']= "='SUBPANEL NTC CHECK'!C18"
    ws6['D32']= "='SUBPANEL NTC CHECK'!C19"
    ws6['D33']= "='SUBPANEL NTC CHECK'!C20"
    ws6['D34']= "='SUBPANEL NTC CHECK'!C21"
    ws6['D35']= "='SUBPANEL NTC CHECK'!C22"
    ws6['D36']= "='SUBPANEL NTC CHECK'!C23"

    ws6['D37']= "=('SUBPANEL NTC CHECK'!C13 + 'SUBPANEL NTC CHECK'!C14 + 'SUBPANEL NTC CHECK'!C15 + 'SUBPANEL NTC CHECK'!C16 + 'SUBPANEL NTC CHECK'!C17 + 'SUBPANEL NTC CHECK'!C18 + 'SUBPANEL NTC CHECK'!C19 + 'SUBPANEL NTC CHECK'!C20 + 'SUBPANEL NTC CHECK'!C21 + 'SUBPANEL NTC CHECK'!C22 + 'SUBPANEL NTC CHECK'!C23)/11"    

    ws6['C37']= "TP53_OVERALL"
        


    ws6['A27']=sampleid
    ws6['A27'].font= Font(bold=True)
    ws6['A27'].font=Font(size=16)

    ws6['A29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['B29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['C29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['D29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['E29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['F29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['G29'].fill= PatternFill("solid", fgColor="FFBB00")

    ws6['A1']=sampleid
    ws6['C1']='Patient Analysis Summary Sheet-CRM'

    ws6['A5'] = sampleid
    ws6['B5']="='Patient demographics'!E2"
    ws6['C5']="='Patient demographics'!F2"
    ws6['D5']= referral
    ws6['E5']= "='Patient demographics'!G2"
    ws6['F5']= "='Patient demographics'!H2"
    ws6['E8']="='Patient demographics'!I2"
    ws6['F8']="='Patient demographics'!J2"
    ws6['G8']="='Patient demographics'!K2"
    ws6['H8']="='Patient demographics'!L2"
    ws6['I8']="='Patient demographics'!M2"
    ws6['J8']="='Variant_calls'!E4"
    ws6['K8']="='Variant_calls'!E7"

    ws6['G5']="='Variant_calls'!G4"
    ws6['H5']="='Variant_calls'!G7"

    ws9['J4']="NTC check 1"
    ws9['J5']="NTC check 2"

    ws2['B4']= sampleid
    ws2['B7']="='Patient demographics'!E2"
    ws2['E4']="='Subpanel NTC check'!K4"
    ws2['E7']="='Subpanel NTC check'!K5"
    ws2['K4']=worksheet


    #change widths of columns
    ws6.column_dimensions['C'].width=40
    ws6.column_dimensions['A'].width=40
    ws6.column_dimensions['B'].width=15
    ws6.column_dimensions['D'].width=10
    ws6.column_dimensions['E'].width=20
    ws6.column_dimensions['F'].width=20
    ws6.column_dimensions['G'].width=20
    ws6.column_dimensions['H'].width=20
    ws9.column_dimensions['A'].width=60
    ws9.column_dimensions['B'].width=15
    ws9.column_dimensions['C'].width=25
    ws9.column_dimensions['D'].width=20
    ws9.column_dimensions['E'].width=15
    ws9.column_dimensions['F'].width=15
    ws2.row_dimensions[9].height=40
    ws2.row_dimensions[61].height=40

    ws2.column_dimensions['C'].width=20
    ws2.column_dimensions['D'].width=53
    ws2.column_dimensions['E'].width=20
    ws2.column_dimensions['H'].width=20
    ws2.column_dimensions['I'].width=20
    ws2.column_dimensions['J'].width=20
    ws2.column_dimensions['K'].width=20
    ws2.column_dimensions['L'].width=20
    ws2.column_dimensions['M'].width=20
    ws2.column_dimensions['N'].width=20
    ws2.column_dimensions['O'].width=20
    ws2.column_dimensions['P'].width=10
    ws2.column_dimensions['Q'].width=30
    ws2.column_dimensions['R'].width=33
    ws2.column_dimensions['S'].width=33
    ws2.column_dimensions['T'].width=40
    ws2.column_dimensions['U'].width=40
    ws2.column_dimensions['V'].width=40

    ws4.column_dimensions['B'].width=20
    ws4.column_dimensions['C'].width=20
    ws4.column_dimensions['D'].width=20
    ws4.column_dimensions['E'].width=20
    ws4.column_dimensions['F'].width=20
    ws4.column_dimensions['G'].width=20
    ws4.column_dimensions['H'].width=20
    ws4.column_dimensions['I'].width=20
    ws4.column_dimensions['J'].width=20
    ws4.column_dimensions['K'].width=20
    ws4.column_dimensions['L'].width=20
    ws4.column_dimensions['M'].width=25
    ws4.column_dimensions['N'].width=25

    ws1.column_dimensions['A'].width=20
    ws1.column_dimensions['B'].width=20
    ws1.column_dimensions['C'].width=20
    ws1.column_dimensions['D'].width=20
    ws1.column_dimensions['E'].width=20
    ws1.column_dimensions['F'].width=20
    ws1.column_dimensions['G'].width=20
    ws1.column_dimensions['H'].width=20
    ws1.column_dimensions['I'].width=20
    ws1.column_dimensions['J'].width=20
    ws1.column_dimensions['K'].width=20
    ws1.column_dimensions['L'].width=20
    ws1.column_dimensions['M'].width=20
    ws1.column_dimensions['N'].width=20
    ws1.column_dimensions['O'].width=20
    ws1.column_dimensions['P'].width=20
    ws1.column_dimensions['Q'].width=20
    ws1.column_dimensions['R'].width=20
    ws1.column_dimensions['S'].width=20
    ws1.column_dimensions['T'].width=20
    ws1.column_dimensions['U'].width=20
    ws1.column_dimensions['V'].width=20
    ws1.column_dimensions['W'].width=20

    ws7.column_dimensions['A'].width=20
    ws7.column_dimensions['B'].width=20
    ws7.column_dimensions['C'].width=20
    ws7.column_dimensions['D'].width=20
    ws7.column_dimensions['E'].width=20
    ws7.column_dimensions['F'].width=20
    ws7.column_dimensions['G'].width=20
    ws7.column_dimensions['H'].width=20
    ws7.column_dimensions['I'].width=20
    ws7.column_dimensions['J'].width=20
    ws7.column_dimensions['K'].width=20
    ws7.column_dimensions['L'].width=20

    ws6['A4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['A5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['E7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['J7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['K7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['J8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['K8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['A12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['A29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['A13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))


    ws6['A17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['E30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['F30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['G30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['A4'].font= Font(bold=True)
    ws6['B4'].font= Font(bold=True)
    ws6['C4'].font= Font(bold=True)
    ws6['D4'].font= Font(bold=True)
    ws6['E4'].font= Font(bold=True)
    ws6['F4'].font= Font(bold=True)
    ws6['G4'].font= Font(bold=True)
    ws6['H4'].font= Font(bold=True)
    ws6['I4'].font= Font(bold=True)

    ws6['E7'].font= Font(bold=True)
    ws6['F7'].font= Font(bold=True)
    ws6['G7'].font= Font(bold=True)
    ws6['H7'].font= Font(bold=True)
    ws6['I7'].font= Font(bold=True)
    ws6['J7'].font= Font(bold=True)
    ws6['K7'].font= Font(bold=True)

    ws6['A1'].font= Font(bold=True)
    ws6['C1'].font= Font(bold=True)
    ws6['A3'].font= Font(bold=True)
    ws6['A9'].font= Font(bold=True)
    ws6['A11'].font= Font(bold=True)

    ws6['H46'].font= Font(bold=True)
    ws6['H47'].font= Font(bold=True)

    ws6['A1'].font=Font(size=16)
    ws6['C1'].font=Font(size=16)
    ws6['H1'].font=Font(size=16)

    ws6['A1'].border=Border(left=Side(border_style=BORDER_THICK), right=Side(border_style=BORDER_THICK), top=Side(border_style=BORDER_THICK), bottom=Side(border_style=BORDER_THICK))
    ws6['C1'].border=Border(left=Side(border_style=BORDER_THICK), right=Side(border_style=BORDER_THICK), top=Side(border_style=BORDER_THICK), bottom=Side(border_style=BORDER_THICK))

    ws1['A1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['B1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['C1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['D1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['E1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['F1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['G1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['H1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['I1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['J1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['K1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['L1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['M1'].fill= PatternFill("solid", fgColor="DCDCDC")

    ws1['A1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['B1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['C1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['D1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['E1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['F1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['G1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['H1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['I1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['J1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['K1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['L1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['M1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws1['A1'].font= Font(bold=True)
    ws1['B1'].font= Font(bold=True)
    ws1['C1'].font= Font(bold=True)
    ws1['D1'].font= Font(bold=True)
    ws1['E1'].font= Font(bold=True)
    ws1['F1'].font= Font(bold=True)
    ws1['G1'].font= Font(bold=True)
    ws1['H1'].font= Font(bold=True)
    ws1['I1'].font= Font(bold=True)
    ws1['J1'].font= Font(bold=True)
    ws1['K1'].font= Font(bold=True)
    ws1['L1'].font= Font(bold=True)
    ws1['M1'].font= Font(bold=True)

    ws2['W9']= "Y/N"
 
    ws2['A9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['B9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['C9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['D9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['E9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['F9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['G9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['H9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['I9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['J9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['K9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['L9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['M9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['N9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['O9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['Q9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['R9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['S9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['T9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['U9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['V9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['W9'].fill= PatternFill("solid", fgColor="DCDCDC") 

    ws2['A9'].font= Font(bold=True)
    ws2['B9'].font= Font(bold=True)
    ws2['C9'].font= Font(bold=True)
    ws2['D9'].font= Font(bold=True)
    ws2['E9'].font= Font(bold=True)
    ws2['F9'].font= Font(bold=True)
    ws2['G9'].font= Font(bold=True)
    ws2['H9'].font= Font(bold=True)
    ws2['I9'].font= Font(bold=True)
    ws2['J9'].font= Font(bold=True)
    ws2['K9'].font= Font(bold=True)
    ws2['L9'].font= Font(bold=True)
    ws2['M9'].font= Font(bold=True)
    ws2['N9'].font= Font(bold=True)
    ws2['O9'].font= Font(bold=True)
    ws2['P9'].font= Font(bold=True)
    ws2['Q9'].font= Font(bold=True)
    ws2['R9'].font= Font(bold=True)
    ws2['S9'].font= Font(bold=True)
    ws2['T9'].font= Font(bold=True)
    ws2['U9'].font= Font(bold=True)
    ws2['V9'].font= Font(bold=True)
    ws2['W9'].font= Font(bold=True)

    ws2['B3'].font= Font(bold=True)
    ws2['B6'].font= Font(bold=True)
    ws2['E3'].font= Font(bold=True)
    ws2['E6'].font= Font(bold=True)
    ws2['G3'].font= Font(bold=True)
    ws2['G6'].font= Font(bold=True)
    ws2['K3'].font= Font(bold=True)
    ws2['K6'].font= Font(bold=True)
    ws2['I3'].font= Font(bold=True)

    ws2['A61'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['B61'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['C61'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['D61'].fill= PatternFill("solid", fgColor="DCDCDC")


    ws2['A61'].font= Font(bold=True)
    ws2['B61'].font= Font(bold=True)
    ws2['C61'].font= Font(bold=True)
    ws2['D61'].font= Font(bold=True)


    ws2['A9'].border=Border(left=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['B9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['C9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['D9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['E9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['F9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['G9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['H9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))    
    ws2['I9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['J9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['K9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['L9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['M9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['N9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['O9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM),right=Side(border_style=BORDER_MEDIUM))
    ws2['Q9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM),left=Side(border_style=BORDER_MEDIUM))
    ws2['R9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['S9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['T9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['U9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['V9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['W9'].border=Border(right=Side(border_style=BORDER_MEDIUM),top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws2['A61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['B61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['C61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['D61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws4['B2'].font= Font(bold=True)
    ws4['C2'].font= Font(bold=True)
    ws4['D2'].font= Font(bold=True)
    ws4['E2'].font= Font(bold=True)
    ws4['F2'].font= Font(bold=True)
    ws4['G2'].font= Font(bold=True)
    ws4['H2'].font= Font(bold=True)
    ws4['I2'].font= Font(bold=True)
    ws4['J2'].font= Font(bold=True)
    ws4['K2'].font= Font(bold=True)
    ws4['L2'].font= Font(bold=True)
    ws4['M2'].font= Font(bold=True)
    ws4['N2'].font= Font(bold=True)

    ws7['A1'].font= Font(bold=True)
    ws7['B1'].font= Font(bold=True)
    ws7['C1'].font= Font(bold=True)
    ws7['D1'].font= Font(bold=True)
    ws7['E1'].font= Font(bold=True)
    ws7['F1'].font= Font(bold=True)
    ws7['G1'].font= Font(bold=True)
    ws7['H1'].font= Font(bold=True)
    ws7['I1'].font= Font(bold=True)
    ws7['J1'].font= Font(bold=True)
    ws7['K1'].font= Font(bold=True)
    ws7['L1'].font= Font(bold=True)



    wb.save(path+sampleid+'_'+referral+'_CRM.xlsx')



if __name__ == "__main__":

    
    #Insert information
    runid=sys.argv[1]
    sampleid=sys.argv[2]
    worksheet=sys.argv[3]
    referral=sys.argv[4]
    NTC_name=sys.argv[5]

    print(runid)
    print(sampleid)
    print(worksheet)
    print(referral)

    path="/data/results/"+runid+"/NGHS-101X/"


    referral=referral.upper()
    if referral=="FOCUS4":
        referral="FOCUS4"
    else:
        print ("referral not recognised")    
    

    referrals_list=['FOCUS4']

    referral_present=False
    
    for referral_value in referrals_list:
        if (referral==referral_value):
            referral_present=True
 
    num_rows_coverage=0

    if (referral_present==True):
    
        variant_report_NTC=get_variantReport_NTC(referral, path)
 
        variant_report_referral=get_variant_report(referral, path, sampleid)

        variant_report_NTC_2=add_extra_columns_NTC_report(variant_report_NTC, variant_report_referral)

        variant_report_referral_2=expand_variant_report(variant_report_referral, variant_report_NTC_2)

        get_gaps_file(referral, path, sampleid)

        hotspots_coverage=get_hotspots_coverage_file(referral, path, sampleid)

        hotspots_coverage_NTC=get_NTC_hotspots_coverage_file(referral, path)

        hotspots_coverage_2, num_rows_coverage=add_columns_hotspots_coverage(hotspots_coverage, hotspots_coverage_NTC)
   
        variant_report_referral_3=match_polys_and_artefacts(variant_report_referral_2, variant_report_NTC_2)

        add_excel_formulae()

    else:
        print("referral not in referrals_list")
