from openpyxl import Workbook 
import pandas 
from openpyxl.utils.dataframe import dataframe_to_rows 
import os
import xlrd 
import sys
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THIN, BORDER_THICK 
from openpyxl.styles import Font

wb=Workbook()
ws1= wb.create_sheet("Sheet_1")
ws9= wb.create_sheet("Sheet_9")
ws2= wb.create_sheet("Sheet_2")
ws4= wb.create_sheet("Sheet_4")
ws5= wb.create_sheet("Sheet_5")
ws6= wb.create_sheet("Sheet_6")
ws7= wb.create_sheet("Sheet_7")
ws10= wb.create_sheet("Sheet_10")

#name the tabs
ws1.title="Patient demographics"
ws2.title="Variant_calls"
ws4.title="Mutations and SNPs"
ws5.title="hotspots.gaps"
ws6.title="Report"
ws7.title="NTC variant"
ws9.title="Subpanel NTC check"
ws10.title="Subpanel coverage"


#set the page layout of the report
ws6.page_setup.orientation=ws6.ORIENTATION_LANDSCAPE
ws6.page_setup.paperSize=ws6.PAPERSIZE_A4


#Patient demographics tab table headers
ws1['A1']='Date Received'
ws1['B1']='Date Requested'
ws1['C1']='LABNO'
ws1['D1']='Patient Name'
ws1['E1']='DOB'
ws1['F1']='Reason for referral'
ws1['G1']='NGS Worksheet'
ws1['H1']='NGS Run Name'

#variant calls table headers
ws2['B3']='DNA number'
ws2['B6']='Patient name'

ws2['E3']='NTC check 1'
ws2['E6']='NTC check 2'

ws2['G3']='1st checker name & date'
ws2['G6']='2nd checker name & date'

ws2['K3']='Myeloid worksheet'
ws2['K6']='Somatic Amplicon v1.7.5'
ws2['A8']=" "


#Mutations and SNPs table headers
ws4['B2']='Gene'
ws4['C2']='Exon/Intron'
ws4['D2']='HGVS c'
ws4['E2']='HGVS p'
ws4['F2']='Allele Frequency'
ws4['G2']='Quality'
ws4['H2']='Depth'
ws4['I2']='Classification'
ws4['J2']='Transcript'
ws4['K2']='Variant'
ws4['L2']='Position'
ws4['M2']='Conclusion 1st checker'
ws4['N2']='Conclusion 2nd checker'


#Report headers
ws6['A3']='Patient information'
ws6['A4']='Lab number'
ws6['B4']='Patient Name'
ws6['C4']='Reason for referral'
ws6['D4']='Panel run'

ws6['E7']='NGS wks'
ws6['H7']='NTC check 1'
ws6['I7']='NTC check 2'

ws6['A4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['B4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['C4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['D4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['E7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['F7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['G7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['H7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['I7'].fill= PatternFill("solid", fgColor="00CCFFFF")

ws6['H46'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['H47'].fill= PatternFill("solid", fgColor="00CCFFFF")

ws6['A11']='Confirmed variant calls'
ws6['A12']='Gene'
ws6['B12']='Exon'
ws6['C12']='Variant'
ws6['D12']='HGVS c.'
ws6['E12']='HGVS p.'
ws6['F12']='Allele frequency'
ws6['G12']='Conclusion 1st checker'
ws6['H12']='Conclusion 2nd checker'

ws6['A12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['B12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['C12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['D12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['E12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['F12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['G12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['H12'].fill= PatternFill("solid", fgColor="009999FF")

    


def get_variantReport_NTC(referral, path): 
    '''
    Fill out the NTC variants tab using the relevant variant report
    '''

    if(os.stat(path+"NTC/hotspot_variants/"+runid+"_NTC_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report_NTC=pandas.read_csv(path+"NTC/hotspot_variants/"+runid+"_NTC_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(path+"NTC/hotspot_variants/"+runid+"_NTC_"+referral+"_VariantReport.txt")
    else:
        variant_report_NTC=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "Preferred","dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
 
    variant_report_NTC_2=pandas.DataFrame(variant_report_NTC)

    #Sort by preferred transcript and filter columns

    variant_report_NTC_3=variant_report_NTC_2[variant_report_NTC_2.Preferred!=False]
    variant_report_NTC_4= variant_report_NTC_3.iloc[:,[23,29,25,26,2,5,3,6,24,1]]
    return (variant_report_NTC_4)




def get_variant_report(referral, path, sampleid):

    '''
    Open the relevant variant file to append to the variant calls tab.
    '''

    if(os.stat(path+ sampleid +"/hotspot_variants/"+runid+"_"+sampleid+"_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report=pandas.read_csv(path+sampleid+"/hotspot_variants/"+runid+"_" +sampleid+"_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(runid+"_"+sampleid+"_"+referral+"_VariantReport.txt")
    else:
        variant_report=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "Preferred", "dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
            

    #sort dataframe by preferred transcript an filter out certain columns
    variant_report_2=pandas.DataFrame(variant_report)
    variant_report_3=variant_report_2[variant_report_2.Preferred!=False]
    variant_report_4= variant_report_3.iloc[:,[23,29,25,26,2,5,3,6,24,1]]
       
    #Add position column in variant calls tab by splitting the variant column 
    variant_list=[]    
    for variant in variant_report_4['Variant']:
        variant_2=[]
        value=0
        for char in variant:
            if(value==0):
                if (char.isdigit()==True or char=="X" or char==":"):
                    variant_2.append(char)
                else:
                    value=1
        variant_3=''.join(variant_2)
        variant_list.append(variant_3)
        
    
    variant_report_4['Position']=variant_list


    return (variant_report_4)





def add_extra_columns_NTC_report(variant_report_NTC_4, variant_report_4):
    
    '''
    Add 'Present in sample' and 'variant allele calls' columns to NTC variant table
    '''

    num_rows_NTC=variant_report_NTC_4.shape[0]
    num_rows_variant_report=variant_report_4.shape[0]

    variant_in_sample=[]
    row=0
    while (row<num_rows_NTC):
        row2=0
        present='NO'
        while (row2<num_rows_variant_report):
            if (variant_report_4.iloc[row2,9]==variant_report_NTC_4.iloc[row,9]):
                present='YES'
            row2=row2+1
        variant_in_sample.append(present)
        row=row+1
     
   
    variant_report_NTC_4['Present in sample']=variant_in_sample

    variant_allele_calls=[]
    row=0
    num_rows_NTC=variant_report_NTC_4.shape[0]
    while(row<num_rows_NTC):
        variant_report_NTC_4.iloc[row,6]=int(variant_report_NTC_4.iloc[row,6])
        allele_call= variant_report_NTC_4.iloc[row,4]*variant_report_NTC_4.iloc[row,6]
        variant_allele_calls.append(allele_call)
        row=row+1

    variant_report_NTC_4['Variant allele calls']=variant_allele_calls    

    for row in dataframe_to_rows(variant_report_NTC_4, header=True, index=False):
        ws7.append(row)

    return (variant_report_NTC_4)






def expand_variant_report(variant_report_4, variant_report_NTC_4):

    '''
    create the extra table at the side of the variant_calls tab
    '''

    detection_threshold=[]
    variant_in_NTC=[]


    row=0
    num_rows_variant_report=variant_report_4.shape[0]


    while (row<num_rows_variant_report):
        variant_report_4.iloc[row,6]= int(variant_report_4.iloc[row,6])
        if (variant_report_4.iloc[row,6]<=500):
            value_2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            value_2=str(value_2)
        else:
            value_2=0
        detection_threshold.append(value_2)
        row2=0
        num_rows_NTC=variant_report_NTC_4.shape[0]
        present='NO'
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row,9]==variant_report_NTC_4.iloc[row2,9]):
                present='YES'
            row2=row2+1
        variant_in_NTC.append(present)
        row=row+1


    variant_report_4["Conclusion 1st checker"]=""
    variant_report_4["QC"]=""
    variant_report_4["Conclusion 2nd checker"]=""
    variant_report_4["QC "]=""
    variant_report_4[""]=""
    variant_report_4["Previously classified"]=""
    variant_report_4["Evidence"]=""

    variant_report_4["Detection threshold based on depth"]=detection_threshold
    variant_report_4["Is variant present in NTC "]=variant_in_NTC


    return (variant_report_4)





def get_gaps_file(referral, path, sampleid):
 
    '''
    Open the relevant gap file to append to the end of the mutations and snps tab. If the gap file is empty, write 'no gaps'.
    '''

    if(os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_" +sampleid+"_"+referral+".gaps").st_size==0):
        ws5['A1']= 'No gaps'
        bedfile=""
    if (os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_" +sampleid+"_"+referral+".gaps").st_size!=0):
        bedfile=pandas.read_csv(path+ sampleid+"/hotspot_coverage/"+runid+"_" +sampleid+"_"+referral+".gaps", sep="\t")
        for row in dataframe_to_rows(bedfile, header=True, index=False):
            ws5.append(row)



    ws6['H46']="Analysed by:"
    ws6['H47']="Checked by:"


    return (bedfile)




def get_hotspots_coverage_file(referral, path, sampleid):

    '''
    Open the relevant coverage file to append to the end of the mutations and snps tab. If the coverage file is empty, write 'No hotspots'.
    '''

    if(os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_"+referral+".coverage").st_size==0):
        ws9['A1']= 'No hotspots'
    if (os.stat(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_"+referral+".coverage").st_size!=0):
        Coverage=pandas.read_csv(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_"+referral+".coverage", sep="\t")
    
    pandas.set_option('display.max_rows',500)   
    Coverage= Coverage.iloc[:,[3,4,5]]
    return(Coverage)




def get_NTC_hotspots_coverage_file(referral, path):

    '''
    Open the relevant NTC hotspots coverage file.
    '''
       
    if(os.stat(path+ "NTC/hotspot_coverage/"+runid+"_NTC_"+referral+".coverage").st_size==0):
        data= [{'CHR':'NA', 'START':'NA', 'END':'NA', 'META':'NA', 'AVG_DEPTH':'NA', 'PERC_COVERAGE@250':'NA'}]
        NTC_check=pandas.DataFrame(data)
    if (os.stat(path+ "NTC/hotspot_coverage/"+runid+"_NTC_"+referral+".coverage").st_size!=0):
        NTC_check=pandas.read_csv(path+ "NTC/hotspot_coverage/"+runid+"_NTC_"+referral+".coverage", sep="\t")
    
    return(NTC_check)




def add_columns_hotspots_coverage(Coverage, NTC_check):

    #Add percentage NTC and subpanel columns to the Coverage table
    
    Coverage['NTC_AVG_Depth']=""
    Coverage['%NTC']=""

    num_rows_NTC= NTC_check.shape[0]
    num_rows_sample= Coverage.shape[0]

    row1=0

    while(row1<num_rows_sample):
        row2=0
        while (row2<num_rows_NTC):
            if(Coverage.iloc[row1,0] == NTC_check.iloc[row2,3]):
                Coverage.iloc[row1,3]= NTC_check.iloc[row2,4]
            row2=row2+1
        row1=row1+1

    Coverage['%NTC']=Coverage['NTC_AVG_Depth']/Coverage['AVG_DEPTH']
    Coverage['%NTC']=Coverage['%NTC'].apply(lambda x: x*100)

    for row in dataframe_to_rows(Coverage, header=True, index=False):
        ws9.append(row)

    num_rows_coverage=Coverage.shape[0]
    row =0
    while (row< num_rows_coverage):
        row_spreadsheet=row+2
        row_spreadsheet_2=str(row_spreadsheet)
        if (Coverage.iloc[row,4]>10):
            ws9['E'+row_spreadsheet_2].fill= PatternFill("solid", fgColor="FFBB00")
        row=row+1



    Coverage['PERC_COVERAGE@500']=Coverage['PERC_COVERAGE@500'].apply(lambda x: int(x))


    Coverage_list=Coverage['PERC_COVERAGE@500']
    less_than_500_coverage_list=[]

    zero_coverage_list=[]

    num_rows_coverage=Coverage.shape[0]
    a=0 
    while (a<num_rows_coverage):
        if (Coverage.iloc[a,2]==0):
            zero_coverage_list.append(Coverage.iloc[a,0])
        if (Coverage.iloc[a,2]<100):
            less_than_500_coverage_list.append(Coverage.iloc[a,0])
            ws9['C'+str(a+2)].fill= PatternFill("solid", fgColor="FFBB00")
        a=a+1

    zero_coverage_list_string=", ".join(zero_coverage_list)
    less_than_500_coverage_list_string=", ".join(less_than_500_coverage_list)


    ws6['A44']=zero_coverage_list_string
    ws6['A49']=less_than_500_coverage_list_string


    return (Coverage, num_rows_coverage)




def get_subpanel_coverage(referral, path, sampleid):

    #Add coverage table
    if(os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_coverage.txt").st_size==0):
        ws10['A1']= 'No coverage'
        Coverage=""
        Coverage_2=""
    if (os.stat(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_coverage.txt").st_size!=0):
        Coverage=pandas.read_csv(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_coverage.txt", sep="\t")

        s=Coverage['FEATURE'].apply(lambda x: x.split('_'))
        Coverage['Referral']=s.apply(lambda x:x[5])

        Coverage_2=Coverage[Coverage.Referral==referral]
 
        for row in dataframe_to_rows(Coverage_2, header=True, index=False):
            ws10.append(row)


    return(Coverage_2)




def match_polys_and_artefacts(variant_report_4, variant_report_NTC_4):

    '''
    Extract the relevant information from the haem artefacts list by matching the variant name with the ones in the variant report table
    '''

    poly_and_Artefact_list=pandas.read_excel("/data/temp/artefacts_lists/HAEM_VARIANT_LIST.xlsx", sheet_name="Benign")
    poly_and_Artefact_list_2=pandas.DataFrame(poly_and_Artefact_list)
    variant_spreadsheet=pandas.read_excel("/data/temp/artefacts_lists/HAEM_VARIANT_LIST.xlsx", sheet_name="Known variants")

    num_rows_variant_report=variant_report_4.shape[0]
    num_rows_poly_artefact=poly_and_Artefact_list_2.shape[0]


    #Fill the conclusion columns using the relevant column in the Poly and Artefact spreadsheet
    row1=0
    while (row1<num_rows_variant_report):
        row2=0
        while(row2<num_rows_poly_artefact):
            if (poly_and_Artefact_list_2.iloc[row2,10]==variant_report_4.iloc[row1,9]):
                variant_report_4.iloc[row1,11]= poly_and_Artefact_list_2.iloc[row2,11]
                variant_report_4.iloc[row1,13]= poly_and_Artefact_list_2.iloc[row2,11]
            row2=row2+1
        row1=row1+1

   
    #fill second table of variant-calls tab using the conclusion column of the first table
    row3=0
    while (row3<num_rows_variant_report):
        if (variant_report_4.iloc[row3,11]=='Known Artefact'):
            variant_report_4.iloc[row3,12]=3
            variant_report_4.iloc[row3,14]=3
        if (variant_report_4.iloc[row3,11]=='Known Polymorphism'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc[row3,14]=1
        if (variant_report_4.iloc[row3,11]=='WT'):
            variant_report_4.iloc[row3,12]=3
            variant_report_4.iloc[row3,14]=3
        if (variant_report_4.iloc[row3,11]=='Genuine'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc[row3,14]=1
        if (variant_report_4.iloc[row3,11]=='SNP'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc[row3,14]=1
														
        row3=row3+1
 

    #Match variants to the variants list to determine what its classification was before
    num_rows_variant_spreadsheet=variant_spreadsheet.shape[0]
    row1=0
    while (row1<num_rows_variant_report):
        row2=0
        while(row2<num_rows_variant_spreadsheet):
            if (variant_spreadsheet.iloc[row2,10]==variant_report_4.iloc[row1,9]):
                variant_report_4.iloc[row1,16]= variant_spreadsheet.iloc[row2,11]
                variant_report_4.iloc[row1,17]=variant_spreadsheet.iloc[row2,12]
            row2=row2+1
        row1=row1+1




    #Add extra columns to the variant report table to determine level of NTC contamination

    
    variant_report_4["#of mutant reads in patient sample "]=""
    variant_report_4["#of mutant reads in NTC if present "]=""
    variant_report_4["Is the NTC contamination significant?"]=""

    num_rows_NTC= variant_report_NTC_4.shape[0]
    row=0


    while (row<num_rows_variant_report):
        if variant_report_4.iloc[row,16]=="YES":
            variant_report_4.iloc[row,6]=float(variant_report_4.iloc[row,6])
            variant_report_4.iloc[row,4]=float(variant_report_4.iloc[row,4])
            value2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            variant_report_4.iloc[row,17]=value2
        row2=0
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row,9]==variant_report_NTC_4.iloc[row2,9]):
                variant_report_4.iloc[row,18]=variant_report_NTC_4.iloc[row2,11]
                variant_report_4.iloc[row,19]=variant_report_4.iloc[row,18]/variant_report_4.iloc[row,17]
            row2=row2+1
        row=row+1


    #Add variant report dataframes to the excel workbook

    variant_report_4=variant_report_4[variant_report_4.Frequency>=0.045]

    for row in dataframe_to_rows(variant_report_4, header=True, index=False):
        ws2.append(row)


    variant_report_5= variant_report_4.iloc[:,[0,1,2]]
    variant_report_5['Comments/Notes/evidence:how conclusion was reached']=""


    row=0

    num_rows_variant_report=variant_report_4.shape[0]

    while (row<num_rows_variant_report):
        if (variant_report_4.iloc[row,11]=='Known artefact'):
            variant_report_5.iloc[row,3]='On artefact list'
        if (variant_report_4.iloc[row,11]=='Known Poly'):
            variant_report_5.iloc[row,3]='On Poly list'
        if (variant_report_4.iloc[row,11]=='WT'):
            variant_report_5.iloc[row,3]='SNP in Ref.Seq'
        row=row+1

    ws2['A60']=" "



    #add dataframe to variant calls tab

    for row in dataframe_to_rows(variant_report_5, header=True, index=False):
        ws2.append(row)

    return(variant_report_4)



def add_excel_formulae():

    #add excel formulae to the spreadsheets to enable automation after program has finished

    ws6['A13']= "='Mutations and SNPS'!B3"
    ws6['B13']= "='Mutations and SNPS'!C3"
    ws6['C13']= "='Mutations and SNPS'!K3"
    ws6['D13']= "='Mutations and SNPS'!D3"
    ws6['E13']= "='Mutations and SNPS'!E3"
    ws6['F13']= "='Mutations and SNPS'!F3"
    ws6['G13']= "='Mutations and SNPS'!M3"
    ws6['H13']= "='Mutations and SNPS'!N3"

    ws6['A14']= "='Mutations and SNPS'!B4"
    ws6['B14']= "='Mutations and SNPS'!C4"
    ws6['C14']= "='Mutations and SNPS'!K4"
    ws6['D14']= "='Mutations and SNPS'!D4"
    ws6['E14']= "='Mutations and SNPS'!E4"
    ws6['F14']= "='Mutations and SNPS'!F4"
    ws6['G14']= "='Mutations and SNPS'!M4"
    ws6['H14']= "='Mutations and SNPS'!N4"

    ws6['A15']= "='Mutations and SNPS'!B5"
    ws6['B15']= "='Mutations and SNPS'!C5"
    ws6['C15']= "='Mutations and SNPS'!K5"
    ws6['D15']= "='Mutations and SNPS'!D5"
    ws6['E15']= "='Mutations and SNPS'!E5"
    ws6['F15']= "='Mutations and SNPS'!F5"
    ws6['G15']= "='Mutations and SNPS'!M5"
    ws6['H15']= "='Mutations and SNPS'!N5"

    ws6['A16']= "='Mutations and SNPS'!B6"
    ws6['B16']= "='Mutations and SNPS'!C6"
    ws6['C16']= "='Mutations and SNPS'!K6"
    ws6['D16']= "='Mutations and SNPS'!D6"
    ws6['E16']= "='Mutations and SNPS'!E6"
    ws6['F16']= "='Mutations and SNPS'!F6"
    ws6['G16']= "='Mutations and SNPS'!M6"
    ws6['H16']= "='Mutations and SNPS'!N6"

    ws6['A17']= "='Mutations and SNPS'!B7"
    ws6['B17']= "='Mutations and SNPS'!C7"
    ws6['C17']= "='Mutations and SNPS'!K7"
    ws6['D17']= "='Mutations and SNPS'!D7"
    ws6['E17']= "='Mutations and SNPS'!E7"
    ws6['F17']= "='Mutations and SNPS'!F7"
    ws6['G17']= "='Mutations and SNPS'!M7"
    ws6['H17']= "='Mutations and SNPS'!N7"

    ws6['A18']= "='Mutations and SNPS'!B8"
    ws6['B18']= "='Mutations and SNPS'!C8"
    ws6['C18']= "='Mutations and SNPS'!K8"
    ws6['D18']= "='Mutations and SNPS'!D8"
    ws6['E18']= "='Mutations and SNPS'!E8"
    ws6['F18']= "='Mutations and SNPS'!F8"
    ws6['G18']= "='Mutations and SNPS'!M8"
    ws6['H18']= "='Mutations and SNPS'!N8"

    ws6['A19']= "='Mutations and SNPS'!B9"
    ws6['B19']= "='Mutations and SNPS'!C9"
    ws6['C19']= "='Mutations and SNPS'!K9"
    ws6['D19']= "='Mutations and SNPS'!D9"
    ws6['E19']= "='Mutations and SNPS'!E9"
    ws6['F19']= "='Mutations and SNPS'!F9"
    ws6['G19']= "='Mutations and SNPS'!M9"
    ws6['H19']= "='Mutations and SNPS'!N9"

    ws6['A20']= "='Mutations and SNPS'!B10"
    ws6['B20']= "='Mutations and SNPS'!C10"
    ws6['C20']= "='Mutations and SNPS'!K10"
    ws6['D20']= "='Mutations and SNPS'!D10"
    ws6['E20']= "='Mutations and SNPS'!E10"
    ws6['F20']= "='Mutations and SNPS'!F10"
    ws6['G20']= "='Mutations and SNPS'!M10"
    ws6['H20']= "='Mutations and SNPS'!N10"

    ws6['A21']= "='Mutations and SNPS'!B11"
    ws6['B21']= "='Mutations and SNPS'!C11"
    ws6['C21']= "='Mutations and SNPS'!K11"
    ws6['D21']= "='Mutations and SNPS'!D11"
    ws6['E21']= "='Mutations and SNPS'!E11"
    ws6['F21']= "='Mutations and SNPS'!F11"
    ws6['G21']= "='Mutations and SNPS'!M11"
    ws6['H21']= "='Mutations and SNPS'!N11"

    ws6['A22']= "='Mutations and SNPS'!B12"
    ws6['B22']= "='Mutations and SNPS'!C12"
    ws6['C22']= "='Mutations and SNPS'!K12"
    ws6['D22']= "='Mutations and SNPS'!D12"
    ws6['E22']= "='Mutations and SNPS'!E12"
    ws6['F22']= "='Mutations and SNPS'!F12"
    ws6['G22']= "='Mutations and SNPS'!M12"
    ws6['H22']= "='Mutations and SNPS'!N12"

    ws6['A23']= "='Mutations and SNPS'!B13"
    ws6['B23']= "='Mutations and SNPS'!C13"
    ws6['C23']= "='Mutations and SNPS'!K13"
    ws6['D23']= "='Mutations and SNPS'!D13"
    ws6['E23']= "='Mutations and SNPS'!E13"
    ws6['F23']= "='Mutations and SNPS'!F13"
    ws6['G23']= "='Mutations and SNPS'!M13"
    ws6['H23']= "='Mutations and SNPS'!N13"

    ws6['A24']= "='Mutations and SNPS'!B14"
    ws6['B24']= "='Mutations and SNPS'!C14"
    ws6['C24']= "='Mutations and SNPS'!K14"
    ws6['D24']= "='Mutations and SNPS'!D14"
    ws6['E24']= "='Mutations and SNPS'!E14"
    ws6['F24']= "='Mutations and SNPS'!F14"
    ws6['G24']= "='Mutations and SNPS'!M14"
    ws6['H24']= "='Mutations and SNPS'!N14"

    ws6['A25']= "='Mutations and SNPS'!B15"
    ws6['B25']= "='Mutations and SNPS'!C15"
    ws6['C25']= "='Mutations and SNPS'!K15"
    ws6['D25']= "='Mutations and SNPS'!D15"
    ws6['E25']= "='Mutations and SNPS'!E15"
    ws6['F25']= "='Mutations and SNPS'!F15"
    ws6['G25']= "='Mutations and SNPS'!M15"
    ws6['H25']= "='Mutations and SNPS'!N15"

    ws6['A26']= "='Mutations and SNPS'!B16"
    ws6['B26']= "='Mutations and SNPS'!C16"
    ws6['C26']= "='Mutations and SNPS'!K16"
    ws6['D26']= "='Mutations and SNPS'!D16"
    ws6['E26']= "='Mutations and SNPS'!E16"
    ws6['F26']= "='Mutations and SNPS'!F16"
    ws6['G26']= "='Mutations and SNPS'!M16"
    ws6['H26']= "='Mutations and SNPS'!N16"

    ws6['A27']= "='Mutations and SNPS'!B17"
    ws6['B27']= "='Mutations and SNPS'!C17"
    ws6['C27']= "='Mutations and SNPS'!K17"
    ws6['D27']= "='Mutations and SNPS'!D17"
    ws6['E27']= "='Mutations and SNPS'!E17"
    ws6['F27']= "='Mutations and SNPS'!F17"
    ws6['G27']= "='Mutations and SNPS'!M17"
    ws6['H27']= "='Mutations and SNPS'!N17"
 
    ws6['A28']= "='Mutations and SNPS'!B18"
    ws6['B28']= "='Mutations and SNPS'!C18"
    ws6['C28']= "='Mutations and SNPS'!K18"
    ws6['D28']= "='Mutations and SNPS'!D18"
    ws6['E28']= "='Mutations and SNPS'!E18"
    ws6['F28']= "='Mutations and SNPS'!F18"
    ws6['G28']= "='Mutations and SNPS'!M18"
    ws6['H28']= "='Mutations and SNPS'!N18"

    ws6['A29']= "='Mutations and SNPS'!B19"
    ws6['B29']= "='Mutations and SNPS'!C19"
    ws6['C29']= "='Mutations and SNPS'!K19"
    ws6['D29']= "='Mutations and SNPS'!D19"
    ws6['E29']= "='Mutations and SNPS'!E19"
    ws6['F29']= "='Mutations and SNPS'!F19"
    ws6['G29']= "='Mutations and SNPS'!M19"
    ws6['H29']= "='Mutations and SNPS'!N19"

    ws6['A30']= "='Mutations and SNPS'!B20"
    ws6['B30']= "='Mutations and SNPS'!C20"
    ws6['C30']= "='Mutations and SNPS'!K20"
    ws6['D30']= "='Mutations and SNPS'!D20"
    ws6['E30']= "='Mutations and SNPS'!E20"
    ws6['F30']= "='Mutations and SNPS'!F20"
    ws6['G30']= "='Mutations and SNPS'!M20"
    ws6['H30']= "='Mutations and SNPS'!N20"


    ws6['A31']= "='Mutations and SNPS'!B21"
    ws6['B31']= "='Mutations and SNPS'!C21"
    ws6['C31']= "='Mutations and SNPS'!K21"
    ws6['D31']= "='Mutations and SNPS'!D21"
    ws6['E31']= "='Mutations and SNPS'!E21"
    ws6['F31']= "='Mutations and SNPS'!F21"
    ws6['G31']= "='Mutations and SNPS'!M21"
    ws6['H31']= "='Mutations and SNPS'!N21"

    ws6['A32']= "='Mutations and SNPS'!B22"
    ws6['B32']= "='Mutations and SNPS'!C22"
    ws6['C32']= "='Mutations and SNPS'!K22"
    ws6['D32']= "='Mutations and SNPS'!D22"
    ws6['E32']= "='Mutations and SNPS'!E22"
    ws6['F32']= "='Mutations and SNPS'!F22"
    ws6['G32']= "='Mutations and SNPS'!M22"
    ws6['H32']= "='Mutations and SNPS'!N22"

    ws6['A33']= "='Mutations and SNPS'!B23"
    ws6['B33']= "='Mutations and SNPS'!C23"
    ws6['C33']= "='Mutations and SNPS'!K23"
    ws6['D33']= "='Mutations and SNPS'!D23"
    ws6['E33']= "='Mutations and SNPS'!E23"
    ws6['F33']= "='Mutations and SNPS'!F23"
    ws6['G33']= "='Mutations and SNPS'!M23"
    ws6['H33']= "='Mutations and SNPS'!N23"

    ws6['A34']= "='Mutations and SNPS'!B24"
    ws6['B34']= "='Mutations and SNPS'!C24"
    ws6['C34']= "='Mutations and SNPS'!K24"
    ws6['D34']= "='Mutations and SNPS'!D24"
    ws6['E34']= "='Mutations and SNPS'!E24"
    ws6['F34']= "='Mutations and SNPS'!F24"
    ws6['G34']= "='Mutations and SNPS'!M24"
    ws6['H34']= "='Mutations and SNPS'!N24"

    ws6['A35']= "='Mutations and SNPS'!B25"
    ws6['B35']= "='Mutations and SNPS'!C25"
    ws6['C35']= "='Mutations and SNPS'!K25"
    ws6['D35']= "='Mutations and SNPS'!D25"
    ws6['E35']= "='Mutations and SNPS'!E25"
    ws6['F35']= "='Mutations and SNPS'!F25"
    ws6['G35']= "='Mutations and SNPS'!M25"
    ws6['H35']= "='Mutations and SNPS'!N25"

    ws6['A36']= "='Mutations and SNPS'!B26"
    ws6['B36']= "='Mutations and SNPS'!C26"
    ws6['C36']= "='Mutations and SNPS'!K26"
    ws6['D36']= "='Mutations and SNPS'!D26"
    ws6['E36']= "='Mutations and SNPS'!E26"
    ws6['F36']= "='Mutations and SNPS'!F26"
    ws6['G36']= "='Mutations and SNPS'!M26"
    ws6['H36']= "='Mutations and SNPS'!N26"

    ws6['A37']= "='Mutations and SNPS'!B27"
    ws6['B37']= "='Mutations and SNPS'!C27"
    ws6['C37']= "='Mutations and SNPS'!K27"
    ws6['D37']= "='Mutations and SNPS'!D27"
    ws6['E37']= "='Mutations and SNPS'!E27"
    ws6['F37']= "='Mutations and SNPS'!F27"
    ws6['G37']= "='Mutations and SNPS'!M27"
    ws6['H37']= "='Mutations and SNPS'!N27"

    ws6['A38']= "='Mutations and SNPS'!B28"
    ws6['B38']= "='Mutations and SNPS'!C28"
    ws6['C38']= "='Mutations and SNPS'!K28"
    ws6['D38']= "='Mutations and SNPS'!D28"
    ws6['E38']= "='Mutations and SNPS'!E28"
    ws6['F38']= "='Mutations and SNPS'!F28"
    ws6['G38']= "='Mutations and SNPS'!M28"
    ws6['H38']= "='Mutations and SNPS'!N28"

    ws6['A39']= "='Mutations and SNPS'!B29"
    ws6['B39']= "='Mutations and SNPS'!C29"
    ws6['C39']= "='Mutations and SNPS'!K29"
    ws6['D39']= "='Mutations and SNPS'!D29"
    ws6['E39']= "='Mutations and SNPS'!E29"
    ws6['F39']= "='Mutations and SNPS'!F29"
    ws6['G39']= "='Mutations and SNPS'!M29"
    ws6['H39']= "='Mutations and SNPS'!N29"

    ws6['A40']= "='Mutations and SNPS'!B30"
    ws6['B40']= "='Mutations and SNPS'!C30"
    ws6['C40']= "='Mutations and SNPS'!K30"
    ws6['D40']= "='Mutations and SNPS'!D30"
    ws6['E40']= "='Mutations and SNPS'!E30"
    ws6['F40']= "='Mutations and SNPS'!F30"
    ws6['G40']= "='Mutations and SNPS'!M30"
    ws6['H40']= "='Mutations and SNPS'!N30"




    ws6['A1']=sampleid

    ws6['A5'] = sampleid
    ws6['B5']="='Patient demographics'!D2"
    ws6['C5']= referral
    ws6['D5']= "Myeloid"
    ws6['E8']= worksheet

    ws9['J4']="NTC check 1"
    ws9['J5']="NTC check 2"

    ws2['D1']='VARIANT-CALLS RESULTS'
    ws2['B4']= sampleid
    ws2['B7']="='Patient demographics'!D2"
    ws2['E4']="='Subpanel NTC check'!K4"
    ws2['E7']="='Subpanel NTC check'!K5"
    ws2['K4']=worksheet

    ws6['A43']="Regions with 0% Coverage"
    ws6['A48']="Regions with <500 Depth"


    ws6['H8']="='Subpanel NTC check'!K4"
    ws6['I8']="='Subpanel NTC check'!K5"


    #change widths of columns
    ws6.column_dimensions['C'].width=40
    ws6.column_dimensions['A'].width=40
    ws6.column_dimensions['B'].width=15
    ws6.column_dimensions['D'].width=10
    ws6.column_dimensions['E'].width=20
    ws6.column_dimensions['F'].width=20
    ws6.column_dimensions['G'].width=20
    ws6.column_dimensions['H'].width=20
    ws9.column_dimensions['A'].width=60
    ws9.column_dimensions['B'].width=15
    ws9.column_dimensions['C'].width=25
    ws9.column_dimensions['D'].width=20
    ws9.column_dimensions['E'].width=15
    ws9.column_dimensions['F'].width=15
    ws2.row_dimensions[9].height=40
    ws2.row_dimensions[61].height=40

    ws2.column_dimensions['C'].width=20
    ws2.column_dimensions['D'].width=53
    ws2.column_dimensions['E'].width=20
    ws2.column_dimensions['H'].width=20
    ws2.column_dimensions['I'].width=20
    ws2.column_dimensions['J'].width=20
    ws2.column_dimensions['K'].width=20
    ws2.column_dimensions['L'].width=20
    ws2.column_dimensions['M'].width=20
    ws2.column_dimensions['N'].width=20
    ws2.column_dimensions['O'].width=20
    ws2.column_dimensions['P'].width=40
    ws2.column_dimensions['Q'].width=30
    ws2.column_dimensions['R'].width=33
    ws2.column_dimensions['S'].width=33
    ws2.column_dimensions['T'].width=40

    ws4.column_dimensions['B'].width=20
    ws4.column_dimensions['C'].width=20
    ws4.column_dimensions['D'].width=20
    ws4.column_dimensions['E'].width=20
    ws4.column_dimensions['F'].width=20
    ws4.column_dimensions['G'].width=20
    ws4.column_dimensions['H'].width=20
    ws4.column_dimensions['I'].width=20
    ws4.column_dimensions['J'].width=20
    ws4.column_dimensions['K'].width=20
    ws4.column_dimensions['L'].width=20
    ws4.column_dimensions['M'].width=25
    ws4.column_dimensions['N'].width=25

    ws1.column_dimensions['A'].width=20
    ws1.column_dimensions['B'].width=20
    ws1.column_dimensions['C'].width=20
    ws1.column_dimensions['D'].width=20
    ws1.column_dimensions['E'].width=20
    ws1.column_dimensions['F'].width=20
    ws1.column_dimensions['G'].width=20
    ws1.column_dimensions['H'].width=20

    ws7.column_dimensions['A'].width=20
    ws7.column_dimensions['B'].width=20
    ws7.column_dimensions['C'].width=20
    ws7.column_dimensions['D'].width=20
    ws7.column_dimensions['E'].width=20
    ws7.column_dimensions['F'].width=20
    ws7.column_dimensions['G'].width=20
    ws7.column_dimensions['H'].width=20
    ws7.column_dimensions['I'].width=20
    ws7.column_dimensions['J'].width=20
    ws7.column_dimensions['K'].width=20
    ws7.column_dimensions['L'].width=20


    ws10.column_dimensions['A'].width=50
    ws10.column_dimensions['B'].width=20
    ws10.column_dimensions['C'].width=25
    ws10.column_dimensions['D'].width=20


    ws6['A4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['A5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['E7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['A12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['A29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['A13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))


    ws6['A17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))


    ws6['A27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H27'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H28'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H29'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['B35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 

    ws6['A4'].font= Font(bold=True)
    ws6['B4'].font= Font(bold=True)
    ws6['C4'].font= Font(bold=True)
    ws6['D4'].font= Font(bold=True)
    ws6['E4'].font= Font(bold=True)
    ws6['F4'].font= Font(bold=True)
    ws6['G4'].font= Font(bold=True)
    ws6['H4'].font= Font(bold=True)
    ws6['I4'].font= Font(bold=True)

    ws6['E7'].font= Font(bold=True)
    ws6['F7'].font= Font(bold=True)
    ws6['G7'].font= Font(bold=True)
    ws6['H7'].font= Font(bold=True)
    ws6['I7'].font= Font(bold=True)


    ws6['A1'].font= Font(bold=True)
    ws6['C1'].font= Font(bold=True)
    ws6['A3'].font= Font(bold=True)
    ws6['A9'].font= Font(bold=True)
    ws6['A11'].font= Font(bold=True)


    ws6['A43'].font= Font(bold=True)
    ws6['A48'].font= Font(bold=True)

    ws6['H46'].font= Font(bold=True)
    ws6['H47'].font= Font(bold=True)

    ws6['H1'].font=Font(size=16)

   
    ws1['A1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['B1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['C1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['D1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['E1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['F1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['G1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['H1'].fill= PatternFill("solid", fgColor="DCDCDC")

    ws1['C2']=sampleid
    ws1['F2']=referral
    ws1['G2']=worksheet
    ws1['H2']=runid


    ws1['A1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['B1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['C1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['D1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['E1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['F1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['G1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['H1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))


    ws1['A1'].font= Font(bold=True)
    ws1['B1'].font= Font(bold=True)
    ws1['C1'].font= Font(bold=True)
    ws1['D1'].font= Font(bold=True)
    ws1['E1'].font= Font(bold=True)
    ws1['F1'].font= Font(bold=True)
    ws1['G1'].font= Font(bold=True)
    ws1['H1'].font= Font(bold=True)


    ws2['D1'].fill= PatternFill("solid", fgColor="DCDCDC") 
    ws2['A9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['B9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['C9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['D9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['E9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['F9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['G9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['H9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['I9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['J9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['K9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['L9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['M9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['N9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['O9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['P9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['Q9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['R9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['S9'].fill= PatternFill("solid", fgColor="DCDCDC")

    ws2['T9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['U9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['V9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['W9'].fill= PatternFill("solid", fgColor="DCDCDC")

   
    ws2['D1'].font= Font(bold=True)
    ws2['D1'].font=Font(size=16)
    ws2['A9'].font= Font(bold=True)
    ws2['B9'].font= Font(bold=True)
    ws2['C9'].font= Font(bold=True)
    ws2['D9'].font= Font(bold=True)
    ws2['E9'].font= Font(bold=True)
    ws2['F9'].font= Font(bold=True)
    ws2['G9'].font= Font(bold=True)
    ws2['H9'].font= Font(bold=True)
    ws2['I9'].font= Font(bold=True)
    ws2['J9'].font= Font(bold=True)
    ws2['K9'].font= Font(bold=True)
    ws2['L9'].font= Font(bold=True)
    ws2['M9'].font= Font(bold=True)
    ws2['N9'].font= Font(bold=True)
    ws2['O9'].font= Font(bold=True)
    ws2['P9'].font= Font(bold=True)
    ws2['Q9'].font= Font(bold=True)
    ws2['R9'].font= Font(bold=True)
    ws2['S9'].font= Font(bold=True)
    ws2['T9'].font= Font(bold=True)
    ws2['U9'].font= Font(bold=True)
    ws2['V9'].font= Font(bold=True)
    ws2['W9'].font= Font(bold=True)

    ws2['B3'].font= Font(bold=True)
    ws2['B6'].font= Font(bold=True)
    ws2['E3'].font= Font(bold=True)
    ws2['E6'].font= Font(bold=True)
    ws2['G3'].font= Font(bold=True)
    ws2['G6'].font= Font(bold=True)
    ws2['K3'].font= Font(bold=True)
    ws2['K6'].font= Font(bold=True)
    ws2['I3'].font= Font(bold=True)

    ws2['A61'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['B61'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['C61'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['D61'].fill= PatternFill("solid", fgColor="DCDCDC")


    ws2['A61'].font= Font(bold=True)
    ws2['B61'].font= Font(bold=True)
    ws2['C61'].font= Font(bold=True)
    ws2['D61'].font= Font(bold=True)


    ws2['A9'].border=Border(left=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['B9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['C9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['D9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['E9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['F9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['G9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['H9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))    
    ws2['I9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['J9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['K9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['L9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['M9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['N9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['O9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['P9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['Q9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['R9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['S9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['T9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['U9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['V9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['W9'].border=Border(right=Side(border_style=BORDER_MEDIUM),top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws2['A61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['B61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['C61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['D61'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws4['B2'].font= Font(bold=True)
    ws4['C2'].font= Font(bold=True)
    ws4['D2'].font= Font(bold=True)
    ws4['E2'].font= Font(bold=True)
    ws4['F2'].font= Font(bold=True)
    ws4['G2'].font= Font(bold=True)
    ws4['H2'].font= Font(bold=True)
    ws4['I2'].font= Font(bold=True)
    ws4['J2'].font= Font(bold=True)
    ws4['K2'].font= Font(bold=True)
    ws4['L2'].font= Font(bold=True)
    ws4['M2'].font= Font(bold=True)
    ws4['N2'].font= Font(bold=True)

    ws7['A1'].font= Font(bold=True)
    ws7['B1'].font= Font(bold=True)
    ws7['C1'].font= Font(bold=True)
    ws7['D1'].font= Font(bold=True)
    ws7['E1'].font= Font(bold=True)
    ws7['F1'].font= Font(bold=True)
    ws7['G1'].font= Font(bold=True)
    ws7['H1'].font= Font(bold=True)
    ws7['I1'].font= Font(bold=True)
    ws7['J1'].font= Font(bold=True)
    ws7['K1'].font= Font(bold=True)
    ws7['L1'].font= Font(bold=True)


    wb.save(path+sampleid+"_"+referral+".xlsx")



if __name__ == "__main__":

    
    #Insert information
    runid=sys.argv[1]
    sampleid=sys.argv[2]
    worksheet=sys.argv[3]
    referral=sys.argv[4]

    print(runid)
    print(sampleid)
    print(worksheet)
    print(referral)

    path="/data/results/"+runid+"/TruSightMyeloid/"


    referral=referral.upper()
    if referral=="MYELOID":
        referral="Myeloid"
    elif referral=="CLL":
        referral="CLL"
    else:
        print ("referral not recognised")    
    

    referrals_list=['Myeloid','CLL']

    referral_present=False
    
    for referral_value in referrals_list:
        if (referral==referral_value):
            referral_present=True
 
    num_rows_coverage=0

    if (referral_present==True):
    
        variant_report_NTC=get_variantReport_NTC(referral, path)
 
        variant_report_referral=get_variant_report(referral, path, sampleid)

        variant_report_NTC_2=add_extra_columns_NTC_report(variant_report_NTC, variant_report_referral)

        variant_report_referral_2=expand_variant_report(variant_report_referral, variant_report_NTC_2)

        gaps_file=get_gaps_file(referral, path, sampleid)

        hotspots_coverage=get_hotspots_coverage_file(referral, path, sampleid)

        hotspots_coverage_NTC=get_NTC_hotspots_coverage_file(referral, path)

        hotspots_coverage_2, num_rows_coverage=add_columns_hotspots_coverage(hotspots_coverage, hotspots_coverage_NTC)
   
        subpanel_coverage=get_subpanel_coverage(referral, path, sampleid)

        variant_report_referral_3=match_polys_and_artefacts(variant_report_referral_2, variant_report_NTC_2)

        add_excel_formulae()

    else:
        print("referral not in referrals_list")
