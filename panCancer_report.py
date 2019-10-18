
from openpyxl import Workbook
import pandas
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import xlrd
import sys
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THIN, BORDER_THICK
from openpyxl.styles import Font

wb=Workbook()
ws1= wb.create_sheet("Sheet_1")
ws9= wb.create_sheet("Sheet_9")
ws10= wb.create_sheet("Sheet_10")
ws2= wb.create_sheet("Sheet_2")
ws4= wb.create_sheet("Sheet_4")
ws5= wb.create_sheet("Sheet_5")
ws6= wb.create_sheet("Sheet_6")
ws7= wb.create_sheet("Sheet_7")
ws8= wb.create_sheet("Sheet_8")


#name the tabs
ws1.title="Patient demographics"
ws2.title="Variant_calls"
ws4.title="Mutations and SNPs"
ws5.title="hotspots.gaps"
ws6.title="Report"
ws7.title="NTC variant"
ws8.title="hotspot_cnvs"
ws9.title="Subpanel NTC check"
ws10.title="Subpanel coverage"


#set the page layout of the report
ws6.page_setup.orientation=ws6.ORIENTATION_LANDSCAPE
ws6.page_setup.paperSize=ws6.PAPERSIZE_A4


#Patient demographics tab table headers
ws1['A1']='Date Received'
ws1['B1']='Date Requested'
ws1['C1']='Due Date'
ws1['D1']='LABNO'
ws1['E1']='Patient name'
ws1['F1']='DOB'
ws1['G1']='Reason for referral'
ws1['H1']='Referring Clinician'
ws1['I1']='Hospital'
ws1['J1']='Date reported'
ws1['K1']='TAT'
ws1['L1']='No of days in histo'
ws1['M1']='Block/Slide/DNA'
ws1['N1']='% Tumour'
ws1['O1']='Result'
ws1['P1']='NGS Worksheet'
ws1['Q1']='Qubit DNA conc. (ng/ul)'
ws1['R1']='Total DNA input'
ws1['S1']='Post PCR1 Qubit'
ws1['T1']='Date of NextSeq run'
ws1['U1']='NextSeq run ID'
ws1['V1']='Comments'


#variant calls table headers
ws2['B3']='DNA number'
ws2['B6']='Patient name'

ws2['E3']='NTC check 1'
ws2['E6']='NTC check 2'

ws2['G3']='1st checker name & date'
ws2['G6']='2nd checker name & date'

ws2['K3']='Pan Cancer worksheet'
ws2['K6']='Analysis pipeline:Roche_PanCancer'
ws2['A8']=" "

ws2['I3']= "% Tumour"


#Mutations and SNPs table headers
ws4['B2']='Gene'
ws4['C2']='Exon/Intron'
ws4['D2']='HGVS c'
ws4['E2']='HGVS p'
ws4['F2']='Allele Frequency'
ws4['G2']='Quality'
ws4['H2']='Depth'
ws4['I2']='Classification'
ws4['J2']='Transcript'
ws4['K2']='Variant'
ws4['L2']='Position'
ws4['M2']='Conclusion 1st checker'
ws4['N2']='Conclusion 2nd checker'


#Report headers
ws6['A3']='Patient information'
ws6['A4']='Lab number'
ws6['B4']='Patient Name'
ws6['C4']='Tumour %'
ws6['D4']='Analysis'
ws6['E4']='Qubit[DNA] ng/ul'
ws6['F4']='Dilution (ng/ul)'
ws6['G4']='Post PCR1 Qubit'
ws6['H4']='Due date'

ws6['E7']='NGS wks'
ws6['F7']='NextSeq run ID'
ws6['G7']='NTC check 1'
ws6['H7']='NTC check 2'

ws6['A11']='Confirmed variant calls'
ws6['A12']='Gene'
ws6['B12']='Exon'
ws6['C12']='Variant'
ws6['D12']='HGVS c.'
ws6['E12']='HGVS p.'
ws6['F12']='Allele frequency'
ws6['G12']='Conclusion 1st checker'
ws6['H12']='Conclusion 2nd checker'



def get_variantReport_NTC(referral, path): 
    '''
    Fill out the NTC variants tab using the relevant variant report
    '''

    if(os.stat(path+"NTC/hotspot_variants/NTC_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report_NTC=pandas.read_csv(path+"NTC/hotspot_variants/NTC_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(path+"NTC/hotspot_variants/NTC_"+referral+"_VariantReport.txt")
    else:
        variant_report_NTC=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "Preferred","dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
 

    variant_report_NTC_2=pandas.DataFrame(variant_report_NTC)

    #Sort by preferred transcript and filter columns
    variant_report_NTC_3=variant_report_NTC_2[variant_report_NTC_2.Preferred!=False]
    variant_report_NTC_4= variant_report_NTC_3.iloc[:,[24,30,26,27,3,6,4,7,25,1]]

    return (variant_report_NTC_4)




def get_variant_report(referral, path, sampleid):

    '''
    Open the relevant variant file to append to the variant calls tab.
    '''

    if(os.stat(path+ sampleid +"/hotspot_variants/"+sampleid+"_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report=pandas.read_csv(path+sampleid+"/hotspot_variants/" +sampleid+"_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(sampleid+"_"+referral+"_VariantReport.txt")
    else:
        variant_report=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "Preferred", "dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
            

    #sort dataframe by preferred transcript an filter out certain columns
    variant_report_2=pandas.DataFrame(variant_report)
    variant_report_3=variant_report_2[variant_report_2.Preferred!=False]
    variant_report_4= variant_report_3.iloc[:,[24,30,26,27,3,6,4,7,25,1]]
    
    
    #Add position column in variant calls tab by splitting the variant column 
    variant_list=[]    
    for variant in variant_report_4['Variant']:
        variant_2=[]
        value=0
        for char in variant:
            if(value==0):
                if (char.isdigit()==True or char=="X" or char==":"):
                    variant_2.append(char)
                else:
                    value=1
        variant_3=''.join(variant_2)
        variant_list.append(variant_3)
        

    variant_report_4['Position']=variant_list


    return (variant_report_4)





def add_extra_columns_NTC_report(variant_report_NTC_4, variant_report_4):
    
    '''
    Add 'Present in sample' and 'variant allele calls' columns to NTC variant table
    '''



    num_rows_NTC=variant_report_NTC_4.shape[0]
    num_rows_variant_report=variant_report_4.shape[0]

    variant_in_sample=[]
    row=0
    while (row<num_rows_NTC):
        row2=0
        present='NO'
        while (row2<num_rows_variant_report):
            if (variant_report_4.iloc[row2,9]==variant_report_NTC_4.iloc[row,9]):
                present='YES'
            row2=row2+1
        variant_in_sample.append(present)
        row=row+1
     
   
    variant_report_NTC_4['Present in sample']=variant_in_sample

    variant_allele_calls=[]
    row=0
    num_rows_NTC=variant_report_NTC_4.shape[0]
    while (row<num_rows_NTC):
        variant_report_NTC_4.iloc[row,6]=int(variant_report_NTC_4.iloc[row,6])
        allele_call= variant_report_NTC_4.iloc[row,4]*variant_report_NTC_4.iloc[row,6]
        variant_allele_calls.append(allele_call)
        row=row+1

    variant_report_NTC_4['Variant allele calls']=variant_allele_calls    

    for row in dataframe_to_rows(variant_report_NTC_4, header=True, index=False):
        ws7.append(row)

    return (variant_report_NTC_4)




def expand_variant_report(variant_report_4, variant_report_NTC_4):

    '''
    create the extra table at the side of the variant_calls tab
    '''

    detection_threshold=[]
    variant_in_NTC=[]


    row=0
    num_rows_variant_report=variant_report_4.shape[0]


    while (row<num_rows_variant_report):
        variant_report_4.iloc[row,6]= int(variant_report_4.iloc[row,6])
        if (variant_report_4.iloc[row,6]<=250):
            value_2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            value_2=str(value_2)
        else:
            value_2=0
        detection_threshold.append(value_2)
        row2=0
        num_rows_NTC=variant_report_NTC_4.shape[0]
        present='NO'
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row,9]==variant_report_NTC_4.iloc[row2,9]):
                present='YES'
            row2=row2+1
        variant_in_NTC.append(present)
        row=row+1


    variant_report_4["Conclusion 1st checker"]=""
    variant_report_4["Conclusion 2nd checker"]=""
    variant_report_4["Comments/Notes/evidence:how conclusion was reached "]=""
    variant_report_4[" "]=""


    variant_report_4["Detection threshold based on depth"]=detection_threshold
    variant_report_4["Is variant present in NTC "]=variant_in_NTC


    return (variant_report_4)




def get_gaps_file(referral, path, sampleid, coverage_value):
 
    '''
    Open the relevant gap file to append to the end of the mutations and snps tab. If the gap file is empty, write 'no gaps'.
    '''
    
    ws5["A1"]="Gaps 250x"
    if(os.stat(path+sampleid+"/hotspot_coverage_"+coverage_value+"/" +sampleid+"_"+referral+"_hotspots.gaps").st_size==0):
        if (coverage_value=="250x"):
            ws5['A2']='No gaps'
            bedfile=""
        elif (coverage_value=="135x"):
            ws5["I1"]="Gaps 135x"
            ws5["I2"]="No gaps"
            bedfile=""
    if (os.stat(path+sampleid+"/hotspot_coverage_"+coverage_value+"/" +sampleid+"_"+referral+"_hotspots.gaps").st_size!=0):
        bedfile=pandas.read_csv(path+ sampleid+"/hotspot_coverage_"+coverage_value+"/" +sampleid+"_"+referral+"_hotspots.gaps", sep="\t")
        if (coverage_value=="135x"):
            ws5["I1"]="Gaps 135x"
            bedfile.columns=["Chromosome","start","end", "annotation"]
            list1=bedfile["Chromosome"].tolist()
            list2=bedfile["start"].tolist()
            list3=bedfile["end"].tolist()
            list4=bedfile["annotation"].tolist()

            number=2
            for value in list1:
                number_2=str(number)
                ws5["I"+number_2]=value
                number=number+1
            number=2
            for value in list2:
                number_2=str(number)
                ws5["J"+number_2]=value
                number=number+1
            number=2
            for value in list3:
                number_2=str(number)
                ws5["K"+number_2]=value
                number=number+1
            number=2
            for value in list4:
                number_2=str(number)
                ws5["L"+number_2]=value
                number=number+1
        elif(coverage_value=="250x"):
            for row in dataframe_to_rows(bedfile, header=True, index=False):
                ws5.append(row)
    return (bedfile)



def get_CNV_file(referral, path, sampleid):
 
    '''
    Open the relevant CNV file to append to the end of the mutations and snps tab. If the CNV file is empty, write 'No CNVs'.
    '''

    if(os.stat(path+ sampleid+"/hotspot_cnvs/"+ sampleid+"_"+referral).st_size==0):
        ws8['A1']= 'No CNVs'
    if (os.stat(path+ sampleid+"/hotspot_cnvs/"+sampleid+"_"+referral).st_size!=0):
        gaps=pandas.read_csv(path+sampleid+"/hotspot_cnvs/"+sampleid+"_"+referral, sep="\t")
       
    for row in dataframe_to_rows(gaps, header=True, index=False):
        ws8.append(row)



    ws6['A48']="=hotspots.gaps!D2"
    ws6['A49']="=hotspots.gaps!D3"
    ws6['A50']="=hotspots.gaps!D4"
    ws6['A51']="=hotspots.gaps!D5"
    ws6['A52']="=hotspots.gaps!D6"
    ws6['A53']="=hotspots.gaps!D7"
    ws6['A54']="=hotspots.gaps!D8"
    ws6['A55']="=hotspots.gaps!D9"
    ws6['A56']="=hotspots.gaps!D10"
    ws6['A57']="=hotspots.gaps!D11"
    ws6['A58']="=hotspots.gaps!D12"
    ws6['A59']="=hotspots.gaps!D13"
    ws6['A60']="=hotspots.gaps!D14"
    ws6['A61']="=hotspots.gaps!D15"
    ws6['A62']="=hotspots.gaps!D16"
    ws6['A63']="=hotspots.gaps!D17"
    ws6['A64']="=hotspots.gaps!D18"
    ws6['A65']="=hotspots.gaps!D19"
    ws6['A66']="=hotspots.gaps!D20"
    ws6['A67']="=hotspots.gaps!D21"
    ws6['A68']="=hotspots.gaps!D22"
    ws6['A69']="=hotspots.gaps!D23"
    ws6['A70']="=hotspots.gaps!D24"
    ws6['A71']="=hotspots.gaps!D25"
    ws6['A72']="=hotspots.gaps!D26"

    ws6['B48']="=hotspots.gaps!D27"
    ws6['B49']="=hotspots.gaps!D28"
    ws6['B50']="=hotspots.gaps!D29"
    ws6['B51']="=hotspots.gaps!D30"
    ws6['B52']="=hotspots.gaps!D31"
    ws6['B53']="=hotspots.gaps!D32"
    ws6['B54']="=hotspots.gaps!D33"
    ws6['B55']="=hotspots.gaps!D34"
    ws6['B56']="=hotspots.gaps!D35"
    ws6['B57']="=hotspots.gaps!D36"
    ws6['B58']="=hotspots.gaps!D37"
    ws6['B59']="=hotspots.gaps!D38"
    ws6['B60']="=hotspots.gaps!D39"
    ws6['B61']="=hotspots.gaps!D40"
    ws6['B62']="=hotspots.gaps!D41"
    ws6['B63']="=hotspots.gaps!D42"
    ws6['B64']="=hotspots.gaps!D43"
    ws6['B65']="=hotspots.gaps!D44"
    ws6['B66']="=hotspots.gaps!D45"
    ws6['B67']="=hotspots.gaps!D46"
    ws6['B68']="=hotspots.gaps!D47"
    ws6['B69']="=hotspots.gaps!D48"
    ws6['B70']="=hotspots.gaps!D49"
    ws6['B71']="=hotspots.gaps!D50"
    ws6['B72']="=hotspots.gaps!D51"


    ws6['C48']="=hotspots.gaps!D52"
    ws6['C49']="=hotspots.gaps!D53"
    ws6['C50']="=hotspots.gaps!D54"
    ws6['C51']="=hotspots.gaps!D55"
    ws6['C52']="=hotspots.gaps!D56"
    ws6['C53']="=hotspots.gaps!D57"
    ws6['C54']="=hotspots.gaps!D58"
    ws6['C55']="=hotspots.gaps!D59"
    ws6['C56']="=hotspots.gaps!D60"
    ws6['C57']="=hotspots.gaps!D61"
    ws6['C58']="=hotspots.gaps!D62"
    ws6['C59']="=hotspots.gaps!D63"
    ws6['C60']="=hotspots.gaps!D64"
    ws6['C61']="=hotspots.gaps!D65"
    ws6['C62']="=hotspots.gaps!D66"
    ws6['C63']="=hotspots.gaps!D67"
    ws6['C64']="=hotspots.gaps!D68"
    ws6['C65']="=hotspots.gaps!D69"
    ws6['C66']="=hotspots.gaps!D70"
    ws6['C67']="=hotspots.gaps!D71"
    ws6['C68']="=hotspots.gaps!D72"
    ws6['C69']="=hotspots.gaps!D73"
    ws6['C70']="=hotspots.gaps!D74"
    ws6['C71']="=hotspots.gaps!D75"
    ws6['C72']="=hotspots.gaps!D76"


    ws6['D48']="=hotspots.gaps!L2"
    ws6['D49']="=hotspots.gaps!L3"
    ws6['D50']="=hotspots.gaps!L4"
    ws6['D51']="=hotspots.gaps!L5"
    ws6['D52']="=hotspots.gaps!L6"
    ws6['D53']="=hotspots.gaps!L7"
    ws6['D54']="=hotspots.gaps!L8"
    ws6['D55']="=hotspots.gaps!L9"
    ws6['D56']="=hotspots.gaps!L10"
    ws6['D57']="=hotspots.gaps!L11"
    ws6['D58']="=hotspots.gaps!L12"
    ws6['D59']="=hotspots.gaps!L13"
    ws6['D60']="=hotspots.gaps!L14"
    ws6['D61']="=hotspots.gaps!L15"
    ws6['D62']="=hotspots.gaps!L16"
    ws6['D63']="=hotspots.gaps!L17"
    ws6['D64']="=hotspots.gaps!L18"
    ws6['D65']="=hotspots.gaps!L19"
    ws6['D66']="=hotspots.gaps!L20"
    ws6['D67']="=hotspots.gaps!L21"
    ws6['D68']="=hotspots.gaps!L22"
    ws6['D69']="=hotspots.gaps!L23"
    ws6['D70']="=hotspots.gaps!L24"
    ws6['D71']="=hotspots.gaps!L25"
    ws6['D72']="=hotspots.gaps!L26"

    ws6['E48']="=hotspots.gaps!L27"
    ws6['E49']="=hotspots.gaps!L28"
    ws6['E50']="=hotspots.gaps!L29"
    ws6['E51']="=hotspots.gaps!L30"
    ws6['E52']="=hotspots.gaps!L31"
    ws6['E53']="=hotspots.gaps!L32"
    ws6['E54']="=hotspots.gaps!L33"
    ws6['E55']="=hotspots.gaps!L34"
    ws6['E56']="=hotspots.gaps!L35"
    ws6['E57']="=hotspots.gaps!L36"
    ws6['E58']="=hotspots.gaps!L37"
    ws6['E59']="=hotspots.gaps!L38"
    ws6['E60']="=hotspots.gaps!L39"
    ws6['E61']="=hotspots.gaps!L40"
    ws6['E62']="=hotspots.gaps!L41"
    ws6['E63']="=hotspots.gaps!L42"
    ws6['E64']="=hotspots.gaps!L43"
    ws6['E65']="=hotspots.gaps!L44"
    ws6['E66']="=hotspots.gaps!L45"
    ws6['E67']="=hotspots.gaps!L46"
    ws6['E68']="=hotspots.gaps!L47"
    ws6['E69']="=hotspots.gaps!L48"
    ws6['E70']="=hotspots.gaps!L49"
    ws6['E71']="=hotspots.gaps!L50"
    ws6['E72']="=hotspots.gaps!L51"

    ws6['F48']="=hotspots.gaps!L52"
    ws6['F49']="=hotspots.gaps!L53"
    ws6['F50']="=hotspots.gaps!L54"
    ws6['F51']="=hotspots.gaps!L55"
    ws6['F52']="=hotspots.gaps!L56"
    ws6['F53']="=hotspots.gaps!L57"
    ws6['F54']="=hotspots.gaps!L58"
    ws6['F55']="=hotspots.gaps!L59"
    ws6['F56']="=hotspots.gaps!L60"
    ws6['F57']="=hotspots.gaps!L61"
    ws6['F58']="=hotspots.gaps!L62"
    ws6['F59']="=hotspots.gaps!L63"
    ws6['F60']="=hotspots.gaps!L64"
    ws6['F61']="=hotspots.gaps!L65"
    ws6['F62']="=hotspots.gaps!L66"
    ws6['F63']="=hotspots.gaps!L67"
    ws6['F64']="=hotspots.gaps!L68"
    ws6['F65']="=hotspots.gaps!L69"
    ws6['F66']="=hotspots.gaps!L70"
    ws6['F67']="=hotspots.gaps!L71"
    ws6['F68']="=hotspots.gaps!L72"
    ws6['F69']="=hotspots.gaps!L73"
    ws6['F70']="=hotspots.gaps!L74"
    ws6['F71']="=hotspots.gaps!L75"
    ws6['F72']="=hotspots.gaps!L76"





    ws6['A47']="Gaps in hotspots ROI 250x"
    ws6['D47']="Gaps in hotspots ROI 135x"

    ws6['F33']="Analysed by:"
    ws6['F34']="Checked by:"


    ws6['A75']="Gene"
    ws6['B75']="Chromosome"
    ws6['C75']="log2"

    return (gaps)




def get_hotspots_coverage_file(referral, path, sampleid, coverage_value):

    '''
    Open the relevant coverage file to append to the end of the mutations and snps tab. If the coverage file is empty, write 'No hotspots'.
    '''

    if(os.stat(path+sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_"+referral+"_hotspots.coverage").st_size==0):
        ws9['A1']= 'No hotspots'
    if (os.stat(path+ sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_"+referral+"_hotspots.coverage").st_size!=0):
        Coverage=pandas.read_csv(path+ sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_"+referral+"_hotspots.coverage", sep="\t")
      

    Coverage= Coverage.iloc[:,[3,4,5]]

    return(Coverage)




def get_NTC_hotspots_coverage_file(referral, path, coverage_value):

    '''
    Open the relevant NTC hotspots coverage file.
    '''
       
    if(os.stat(path+ "NTC/hotspot_coverage_"+coverage_value+"/NTC_"+referral+"_hotspots.coverage").st_size==0):
        data= [{'CHR':'NA', 'START':'NA', 'END':'NA', 'META':'NA', 'AVG_DEPTH':'NA', 'PERC_COVERAGE@250':'NA'}]
        NTC_check=pandas.DataFrame(data)
    if (os.stat(path+ "NTC/hotspot_coverage_"+coverage_value+"/NTC_"+referral+"_hotspots.coverage").st_size!=0):
        NTC_check=pandas.read_csv(path+ "NTC/hotspot_coverage_"+coverage_value+"/NTC_"+referral+"_hotspots.coverage", sep="\t")
    
    return(NTC_check)



def add_columns_hotspots_coverage(Coverage, NTC_check):

    if(os.stat(path+sampleid+"/hotspot_coverage_135x/"+sampleid+"_"+referral+"_hotspots.coverage").st_size==0):
        ws9['A1']= 'No hotspots'
    if (os.stat(path+ sampleid+"/hotspot_coverage_135x/"+sampleid+"_"+referral+"_hotspots.coverage").st_size!=0):
        Coverage_135x=pandas.read_csv(path+ sampleid+"/hotspot_coverage_135x/"+sampleid+"_"+referral+"_hotspots.coverage", sep="\t")
        Coverage["PERC_COVERAGE@135"]=Coverage_135x["PERC_COVERAGE@135"]
    #Add percentage NTC and subpanel columns to the Coverage table
    
    Coverage['NTC_AVG_Depth']=""
    Coverage['%NTC']=""

    num_rows_NTC= NTC_check.shape[0]
    num_rows_sample= Coverage.shape[0]

    row1=0

    while(row1<num_rows_sample):
        row2=0
        while (row2<num_rows_NTC):
            if(Coverage.iloc[row1,0] == NTC_check.iloc[row2,3]):
                Coverage.iloc[row1,4]= NTC_check.iloc[row2,4]
            row2=row2+1
        row1=row1+1
    
    Coverage['%NTC']=Coverage['NTC_AVG_Depth']/Coverage['AVG_DEPTH']
    Coverage['%NTC']= Coverage['%NTC']*100
    Coverage['Screen Type']= "Hotspots"

    for row in dataframe_to_rows(Coverage, header=True, index=False):
        ws9.append(row)

    num_rows_coverage=Coverage.shape[0]
    row =0
    while (row< num_rows_coverage):
        row_spreadsheet=row+2
        row_spreadsheet_2=str(row_spreadsheet)
        if (Coverage.iloc[row,4]>10):
            ws9['E'+row_spreadsheet_2].fill= PatternFill("solid", fgColor="FFBB00")
        ws9['G'+row_spreadsheet_2].fill= PatternFill("solid", fgColor="00CCFFFF")
        row=row+1
    
    return (Coverage, num_rows_coverage)





def get_genescreen_coverage_file(referral, path, sampleid,coverage_value):

    '''
    Open the coverage file to append to the mutations and snps tab. If the coverage file is empty, write 'No hotspots'.
    '''
    if(os.stat(path+sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_"+referral+"_genescreen.coverage").st_size==0):
        ws9['A1']= 'No hotspots'
    if (os.stat(path+ sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_"+referral+"_genescreen.coverage").st_size!=0):
        Coverage=pandas.read_csv(path+ sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_"+referral+"_genescreen.coverage", sep="\t")
    
    Coverage= Coverage.iloc[:,[3,4,5]]

    return(Coverage)




def get_NTC_genescreen_coverage_file(referral, path,coverage_value):

    '''
    Open the relevant NTC  genescreen coverage file to append to the Subpanel NTC check tab.
    '''

    if(os.stat(path+ "/NTC/hotspot_coverage_"+coverage_value+"/NTC_"+referral+"_genescreen.coverage").st_size==0):
        data= [{'CHR':'NA', 'START':'NA', 'END':'NA', 'META':'NA', 'AVG_DEPTH':'NA', 'PERC_COVERAGE@250':'NA'}]
        NTC_check=pandas.DataFrame(data)
    if (os.stat(path+ "/NTC/hotspot_coverage_"+coverage_value+"/NTC_"+referral+"_genescreen.coverage").st_size!=0):
        NTC_check=pandas.read_csv(path+ "/NTC/hotspot_coverage_"+coverage_value+"/NTC_"+referral+"_genescreen.coverage", sep="\t")
       

    return (NTC_check)





def add_columns_genescreen_coverage(Coverage, NTC_check, num_rows_coverage):

    '''
    Add NTC and subpanel columns to the Coverage table
    '''

    if(os.stat(path+sampleid+"/hotspot_coverage_135x/"+sampleid+"_"+referral+"_genescreen.coverage").st_size==0):
        ws9['A1']= 'No hotspots'
    if (os.stat(path+ sampleid+"/hotspot_coverage_135x/"+sampleid+"_"+referral+"_genescreen.coverage").st_size!=0):
        Coverage_135x=pandas.read_csv(path+ sampleid+"/hotspot_coverage_135x/"+sampleid+"_"+referral+"_genescreen.coverage", sep="\t")
        Coverage["PERC_COVERAGE@135"]=Coverage_135x["PERC_COVERAGE@135"]

    Coverage['NTC_AVG_Depth']=""
    Coverage['%NTC']=""

    num_rows_NTC= NTC_check.shape[0]
    num_rows_sample= Coverage.shape[0]

    row1=0

    while(row1<num_rows_sample):
        row2=0
        while (row2<num_rows_NTC):
            if(Coverage.iloc[row1,0] == NTC_check.iloc[row2,3]):
                Coverage.iloc[row1,4]= NTC_check.iloc[row2,4]
            row2=row2+1
        row1=row1+1
    
   
    Coverage['%NTC']=Coverage['NTC_AVG_Depth']/Coverage['AVG_DEPTH']
    Coverage['%NTC']= Coverage['%NTC']*100
 
   
    Coverage['Screen Type']= "Gene screen"


    for row in dataframe_to_rows(Coverage, header=False, index=False):
        ws9.append(row)

    num_rows_coverage_2=Coverage.shape[0]
    num_rows_coverage_3=num_rows_coverage +num_rows_coverage_2

    row =0
    while (row< num_rows_coverage_2):
        row_spreadsheet=row+2+num_rows_coverage
        row_spreadsheet_2=str(row_spreadsheet)
        if (Coverage.iloc[row,4]>10):
            ws9['E'+row_spreadsheet_2].fill= PatternFill("solid", fgColor="FFBB00")
        ws9['G'+row_spreadsheet_2].fill= PatternFill("solid", fgColor="009999FF")
        row=row+1


    return(Coverage)




def get_subpanel_coverage(referral, path, sampleid, coverage_value):

    #Add coverage table
    if(os.stat(path+sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_coverage.txt").st_size==0):
        ws10['A1']= 'No coverage'
        Coverage=""
        Coverage_2=""
    if (os.stat(path+ sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_coverage.txt").st_size!=0):
        Coverage=pandas.read_csv(path+ sampleid+"/hotspot_coverage_"+coverage_value+"/"+sampleid+"_coverage.txt", sep="\t")

    if (os.stat(path+ sampleid+"/hotspot_coverage_135x/"+sampleid+"_coverage.txt").st_size!=0):
        Coverage_135x=pandas.read_csv(path+ sampleid+"/hotspot_coverage_135x/"+sampleid+"_coverage.txt", sep="\t")
        Coverage["PERC_COVERAGE@135"]=Coverage_135x["PERC_COVERAGE@135"]

        s=Coverage['FEATURE'].apply(lambda x: x.split('_'))
        Coverage['Referral']=s.apply(lambda x:x[1])

        Coverage_2=Coverage[Coverage.Referral==referral]
 
    for row in dataframe_to_rows(Coverage_2, header=True, index=False):
        ws10.append(row)

    

    ws6['A30']="='Subpanel coverage'!A2"
    ws6['A31']="='Subpanel coverage'!A3"
    ws6['A32']= "='Subpanel coverage'!A4"
    ws6['A33']="='Subpanel coverage'!A5"
    ws6['A34']="='Subpanel coverage'!A6"
    ws6['A35']="='Subpanel coverage'!A7"
    ws6['A36']="='Subpanel coverage'!A8"
    ws6['A37']="='Subpanel coverage'!A9"
    ws6['A38']="='Subpanel coverage'!A10"
    ws6['A39']="='Subpanel coverage'!A11"
    ws6['A40']="='Subpanel coverage'!A12"
    ws6['A41']="='Subpanel coverage'!A13"
    ws6['A42']="='Subpanel coverage'!A14"
    ws6['A43']="='Subpanel coverage'!A15"

    ws6['B30']="='Subpanel coverage'!C2"
    ws6['B31']="='Subpanel coverage'!C3"
    ws6['B32']="='Subpanel coverage'!C4"
    ws6['B33']="='Subpanel coverage'!C5"
    ws6['B34']="='Subpanel coverage'!C6"
    ws6['B35']="='Subpanel coverage'!C7"
    ws6['B36']="='Subpanel coverage'!C8"
    ws6['B37']="='Subpanel coverage'!C9"
    ws6['B38']="='Subpanel coverage'!C10"
    ws6['B39']="='Subpanel coverage'!C11"
    ws6['B40']= "='Subpanel coverage'!C12"
    ws6['B41']="='Subpanel coverage'!C13"
    ws6['B42']="='Subpanel coverage'!C14"
    ws6['B43']="='Subpanel coverage'!C15"

    ws6['C30']="='Subpanel coverage'!D2"
    ws6['C31']="='Subpanel coverage'!D3"
    ws6['C32']="='Subpanel coverage'!D4"
    ws6['C33']="='Subpanel coverage'!D5"
    ws6['C34']="='Subpanel coverage'!D6"
    ws6['C35']="='Subpanel coverage'!D7"
    ws6['C36']="='Subpanel coverage'!D8"
    ws6['C37']="='Subpanel coverage'!D9"
    ws6['C38']="='Subpanel coverage'!D10"
    ws6['C39']="='Subpanel coverage'!D11"
    ws6['C40']= "='Subpanel coverage'!D12"
    ws6['C41']="='Subpanel coverage'!D13"
    ws6['C42']="='Subpanel coverage'!D14"
    ws6['C43']="='Subpanel coverage'!D15"



    return(Coverage_2)




def match_polys_and_artefacts(variant_report_4, variant_report_NTC_4):

    '''
    Extract the relevant information from "PanCancer_Poly and Artefact list.xlsx" by matching the variant name with the ones in the variant report table
    '''

    poly_artefact_dict={}
    poly_and_Artefact_list=pandas.read_excel("/data/temp/artefacts_lists/Pan_Poly_and_Artefact_list.xlsx")
    poly_and_Artefact_list_2=pandas.DataFrame(poly_and_Artefact_list)


    num_rows_variant_report=variant_report_4.shape[0]
    num_rows_poly_artefact=poly_and_Artefact_list_2.shape[0]


    #Fill the conclusion columns using the relevant column in the Poly and Artefact spreadsheet
    row1=0
    while (row1<num_rows_variant_report):
        row2=0
        while(row2<num_rows_poly_artefact):
            if (poly_and_Artefact_list_2.iloc[row2,9]==variant_report_4.iloc[row1,9]):
                variant_report_4.iloc[row1,11]= poly_and_Artefact_list_2.iloc[row2,13]
                variant_report_4.iloc[row1,12]= poly_and_Artefact_list_2.iloc[row2,13]
                if (variant_report_4.iloc[row1,11]=='Known artefact'):
                    variant_report_4.iloc[row1,13]='On artefact list'
                elif (variant_report_4.iloc[row1,11]=='Known Poly'):
                    variant_report_4.iloc[row1,13]='On Poly list'
                elif (variant_report_4.iloc[row1,11]=='WT'):
                    variant_report_4.iloc[row1,13]='SNP in Ref.Seq'
                else:
                    variant_report_4.iloc[row1,13]= ""
            row2=row2+1
        row1=row1+1

 

   #Add extra columns to the variant report table to determine level of NTC contamination

    
    variant_report_4["#of mutant reads in patient sample"]=""
    variant_report_4["#of mutant reads in NTC if present "]=""
    variant_report_4["Is the NTC contamination significant?"]=""


    num_rows_NTC= variant_report_NTC_4.shape[0]
    row=0


    while (row<num_rows_variant_report):
        if variant_report_4.iloc[row,16]=="YES":
            variant_report_4.iloc[row,6]=float(variant_report_4.iloc[row,6])
            variant_report_4.iloc[row,4]=float(variant_report_4.iloc[row,4])
            value2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            variant_report_4.iloc[row,17]=value2
        row2=0
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row, 9]==variant_report_NTC_4.iloc[row2,9]):
                variant_report_4.iloc[row,18]=variant_report_NTC_4.iloc[row2,11]
                variant_report_4.iloc[row,19]=variant_report_4.iloc[row,18]/variant_report_4.iloc[row,17]
            row2=row2+1
        row=row+1


    #Add upper-limit and lower-limit variant report dataframes to the excel workbook
    
    variant_report_4_upper_limit=variant_report_4[variant_report_4.Frequency>0.045]

    for row in dataframe_to_rows(variant_report_4_upper_limit, header=True, index=False):
        ws2.append(row)


    return(variant_report_4)



def add_excel_formulae():

    #add excel formulae to the spreadsheets to enable automation after program has finished

    ws2['I4']= "='Patient demographics'!N2"
     
    ws6['A13']= "='Mutations and SNPS'!B3"
    ws6['B13']= "='Mutations and SNPS'!C3"
    ws6['C13']= "='Mutations and SNPS'!K3"
    ws6['D13']= "='Mutations and SNPS'!D3"
    ws6['E13']= "='Mutations and SNPS'!E3"
    ws6['F13']= "='Mutations and SNPS'!F3"
    ws6['G13']= "='Mutations and SNPS'!M3"
    ws6['H13']= "='Mutations and SNPS'!N3"

    ws6['A14']= "='Mutations and SNPS'!B4"
    ws6['B14']= "='Mutations and SNPS'!C4"
    ws6['C14']= "='Mutations and SNPS'!K4"
    ws6['D14']= "='Mutations and SNPS'!D4"
    ws6['E14']= "='Mutations and SNPS'!E4"
    ws6['F14']= "='Mutations and SNPS'!F4"
    ws6['G14']= "='Mutations and SNPS'!M4"
    ws6['H14']= "='Mutations and SNPS'!N4"

    ws6['A15']= "='Mutations and SNPS'!B5"
    ws6['B15']= "='Mutations and SNPS'!C5"
    ws6['C15']= "='Mutations and SNPS'!K5"
    ws6['D15']= "='Mutations and SNPS'!D5"
    ws6['E15']= "='Mutations and SNPS'!E5"
    ws6['F15']= "='Mutations and SNPS'!F5"
    ws6['G15']= "='Mutations and SNPS'!M5"
    ws6['H15']= "='Mutations and SNPS'!N5"

    ws6['A16']= "='Mutations and SNPS'!B6"
    ws6['B16']= "='Mutations and SNPS'!C6"
    ws6['C16']= "='Mutations and SNPS'!K6"
    ws6['D16']= "='Mutations and SNPS'!D6"
    ws6['E16']= "='Mutations and SNPS'!E6"
    ws6['F16']= "='Mutations and SNPS'!F6"
    ws6['G16']= "='Mutations and SNPS'!M6"
    ws6['H16']= "='Mutations and SNPS'!N6"

    ws6['A17']= "='Mutations and SNPS'!B7"
    ws6['B17']= "='Mutations and SNPS'!C7"
    ws6['C17']= "='Mutations and SNPS'!K7"
    ws6['D17']= "='Mutations and SNPS'!D7"
    ws6['E17']= "='Mutations and SNPS'!E7"
    ws6['F17']= "='Mutations and SNPS'!F7"
    ws6['G17']= "='Mutations and SNPS'!M7"
    ws6['H17']= "='Mutations and SNPS'!N7"

    ws6['A18']= "='Mutations and SNPS'!B8"
    ws6['B18']= "='Mutations and SNPS'!C8"
    ws6['C18']= "='Mutations and SNPS'!K8"
    ws6['D18']= "='Mutations and SNPS'!D8"
    ws6['E18']= "='Mutations and SNPS'!E8"
    ws6['F18']= "='Mutations and SNPS'!F8"
    ws6['G18']= "='Mutations and SNPS'!M8"
    ws6['H18']= "='Mutations and SNPS'!N8"

    ws6['A19']= "='Mutations and SNPS'!B9"
    ws6['B19']= "='Mutations and SNPS'!C9"
    ws6['C19']= "='Mutations and SNPS'!K9"
    ws6['D19']= "='Mutations and SNPS'!D9"
    ws6['E19']= "='Mutations and SNPS'!E9"
    ws6['F19']= "='Mutations and SNPS'!F9"
    ws6['G19']= "='Mutations and SNPS'!M9"
    ws6['H19']= "='Mutations and SNPS'!N9"

    ws6['A20']= "='Mutations and SNPS'!B10"
    ws6['B20']= "='Mutations and SNPS'!C10"
    ws6['C20']= "='Mutations and SNPS'!K10"
    ws6['D20']= "='Mutations and SNPS'!D10"
    ws6['E20']= "='Mutations and SNPS'!E10"
    ws6['F20']= "='Mutations and SNPS'!F10"
    ws6['G20']= "='Mutations and SNPS'!M10"
    ws6['H20']= "='Mutations and SNPS'!N10"

    ws6['A21']= "='Mutations and SNPS'!B11"
    ws6['B21']= "='Mutations and SNPS'!C11"
    ws6['C21']= "='Mutations and SNPS'!K11"
    ws6['D21']= "='Mutations and SNPS'!D11"
    ws6['E21']= "='Mutations and SNPS'!E11"
    ws6['F21']= "='Mutations and SNPS'!F11"
    ws6['G21']= "='Mutations and SNPS'!M11"
    ws6['H21']= "='Mutations and SNPS'!N11"

    ws6['A22']= "='Mutations and SNPS'!B12"
    ws6['B22']= "='Mutations and SNPS'!C12"
    ws6['C22']= "='Mutations and SNPS'!K12"
    ws6['D22']= "='Mutations and SNPS'!D12"
    ws6['E22']= "='Mutations and SNPS'!E12"
    ws6['F22']= "='Mutations and SNPS'!F12"
    ws6['G22']= "='Mutations and SNPS'!M12"
    ws6['H22']= "='Mutations and SNPS'!N12"

    ws6['A23']= "='Mutations and SNPS'!B13"
    ws6['B23']= "='Mutations and SNPS'!C13"
    ws6['C23']= "='Mutations and SNPS'!K13"
    ws6['D23']= "='Mutations and SNPS'!D13"
    ws6['E23']= "='Mutations and SNPS'!E13"
    ws6['F23']= "='Mutations and SNPS'!F13"
    ws6['G23']= "='Mutations and SNPS'!M13"
    ws6['H23']= "='Mutations and SNPS'!N13"

    ws6['A24']= "='Mutations and SNPS'!B14"
    ws6['B24']= "='Mutations and SNPS'!C14"
    ws6['C24']= "='Mutations and SNPS'!K14"
    ws6['D24']= "='Mutations and SNPS'!D14"
    ws6['E24']= "='Mutations and SNPS'!E14"
    ws6['F24']= "='Mutations and SNPS'!F14"
    ws6['G24']= "='Mutations and SNPS'!M14"
    ws6['H24']= "='Mutations and SNPS'!N14"

    ws6['A25']= "='Mutations and SNPS'!B15"
    ws6['B25']= "='Mutations and SNPS'!C15"
    ws6['C25']= "='Mutations and SNPS'!K15"
    ws6['D25']= "='Mutations and SNPS'!D15"
    ws6['E25']= "='Mutations and SNPS'!E15"
    ws6['F25']= "='Mutations and SNPS'!F15"
    ws6['G25']= "='Mutations and SNPS'!M15"
    ws6['H25']= "='Mutations and SNPS'!N15"

    ws6['A26']= "='Mutations and SNPS'!B16"
    ws6['B26']= "='Mutations and SNPS'!C16"
    ws6['C26']= "='Mutations and SNPS'!K16"
    ws6['D26']= "='Mutations and SNPS'!D16"
    ws6['E26']= "='Mutations and SNPS'!E16"
    ws6['F26']= "='Mutations and SNPS'!F16"
    ws6['G26']= "='Mutations and SNPS'!M16"
    ws6['H26']= "='Mutations and SNPS'!N16"

    
    ws6['B29']="Percentage of bases covered to 250x" 

    ws6['D29']="Comments"
    ws6['A29']= sampleid +"_" + referral
    ws6['C29'] = "Percentage of bases covered to 135x"
    ws6['A74']= "CNV results"

    ws6['A27']=sampleid
    ws6['A27'].font= Font(bold=True)
    ws6['A27'].font=Font(size=16)

    ws6['A29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['B29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['C29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['D29'].fill= PatternFill("solid", fgColor="FFBB00")

    ws6['A47'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['B47'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['C47'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['D47'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['E47'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['F47'].fill= PatternFill("solid", fgColor="FFBB00")

    ws6['A75'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['B75'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['C75'].fill= PatternFill("solid", fgColor="FFBB00")


    ws6['A76']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N2&" ")),"",hotspot_cnvs!A2)'
    ws6['A77']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N3&" ")),"",hotspot_cnvs!A3)'
    ws6['A78']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N4&" ")),"",hotspot_cnvs!A4)'
    ws6['A79']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N5&" ")),"",hotspot_cnvs!A5)'
    ws6['A80']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N6&" ")),"",hotspot_cnvs!A6)'
    ws6['A81']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N7&" ")),"",hotspot_cnvs!A7)'
    ws6['A82']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N8&" ")),"",hotspot_cnvs!A8)'
    ws6['A83']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N9&" ")),"",hotspot_cnvs!A9)'
    ws6['A84']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N10&" ")),"",hotspot_cnvs!A10)'

    ws6['B76']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N2&" ")),"",hotspot_cnvs!B2)'
    ws6['B77']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N3&" ")),"",hotspot_cnvs!B3)'
    ws6['B78']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N4&" ")),"",hotspot_cnvs!B4)'
    ws6['B79']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N5&" ")),"",hotspot_cnvs!B5)'
    ws6['B80']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N6&" ")),"",hotspot_cnvs!B6)'  
    ws6['B81']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N7&" ")),"",hotspot_cnvs!B7)'
    ws6['B82']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N8&" ")),"",hotspot_cnvs!B8)'
    ws6['B83']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N9&" ")),"",hotspot_cnvs!B9)'
    ws6['B84']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N10&" ")),"",hotspot_cnvs!B10)'

    ws6['C76']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N2&" ")),"",hotspot_cnvs!E2)'
    ws6['C77']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N3&" ")),"",hotspot_cnvs!E3)'
    ws6['C78']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N4&" ")),"",hotspot_cnvs!E4)'
    ws6['C79']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N5&" ")),"",hotspot_cnvs!E5)'
    ws6['C80']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N6&" ")),"",hotspot_cnvs!E6)'
    ws6['C81']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N7&" ")),"",hotspot_cnvs!E7)'
    ws6['C82']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N8&" ")),"",hotspot_cnvs!E8)'
    ws6['C83']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N9&" ")),"",hotspot_cnvs!E9)'
    ws6['C84']= '=IF(ISERR(SEARCH("Genuine"," "& hotspot_cnvs!N10&" ")),"",hotspot_cnvs!E10)'



    ws6['A1']=sampleid
    ws6['C1']='Patient Analysis Summary Sheet-PanCancer'

    ws6['A5'] = sampleid
    ws6['B5']="='Patient demographics'!E2"
    ws6['C5']="='Patient demographics'!N2"
    ws6['D5']= referral
    ws6['E5']= "='Patient demographics'!Q2"
    ws6['F5']= "='Patient demographics'!R2"
    ws6['G5']= "='Patient demographics'!S2"
    ws6['H5']= "='Patient demographics'!C2"
    ws6['E8']="='Patient demographics'!P2"
    ws6['F8']="='Patient demographics'!W2"
    ws6['G8']="='Patient demographics'!K4"
    ws6['H8']= "='Subpanel NTC check'!K5"
        

    ws9['J4']="NTC check 1"
    ws9['J5']="NTC check 2"

    ws2['B4']= sampleid
    ws2['B7']="='Patient demographics'!E2"
    ws2['E4']="='Subpanel NTC check'!K4"
    ws2['E7']="='Subpanel NTC check'!K5"
    ws2['K4']=worksheet


    #change widths of columns
    ws6.column_dimensions['C'].width=40
    ws6.column_dimensions['A'].width=40
    ws6.column_dimensions['B'].width=15
    ws6.column_dimensions['D'].width=10
    ws6.column_dimensions['E'].width=20
    ws6.column_dimensions['F'].width=20
    ws6.column_dimensions['G'].width=20
    ws6.column_dimensions['H'].width=20
    ws9.column_dimensions['A'].width=60
    ws9.column_dimensions['B'].width=15
    ws9.column_dimensions['C'].width=25
    ws9.column_dimensions['D'].width=20
    ws9.column_dimensions['E'].width=15
    ws9.column_dimensions['F'].width=15
    ws2.row_dimensions[9].height=40
    ws2.row_dimensions[61].height=40

    ws2.column_dimensions['C'].width=20
    ws2.column_dimensions['D'].width=53
    ws2.column_dimensions['E'].width=20
    ws2.column_dimensions['H'].width=20
    ws2.column_dimensions['I'].width=20
    ws2.column_dimensions['J'].width=20
    ws2.column_dimensions['K'].width=20
    ws2.column_dimensions['L'].width=20
    ws2.column_dimensions['M'].width=20
    ws2.column_dimensions['N'].width=20
    ws2.column_dimensions['O'].width=20
    ws2.column_dimensions['P'].width=40
    ws2.column_dimensions['Q'].width=30
    ws2.column_dimensions['R'].width=33
    ws2.column_dimensions['S'].width=33
    ws2.column_dimensions['T'].width=40

    ws4.column_dimensions['B'].width=20
    ws4.column_dimensions['C'].width=20
    ws4.column_dimensions['D'].width=20
    ws4.column_dimensions['E'].width=20
    ws4.column_dimensions['F'].width=20
    ws4.column_dimensions['G'].width=20
    ws4.column_dimensions['H'].width=20
    ws4.column_dimensions['I'].width=20
    ws4.column_dimensions['J'].width=20
    ws4.column_dimensions['K'].width=20
    ws4.column_dimensions['L'].width=20
    ws4.column_dimensions['M'].width=25
    ws4.column_dimensions['N'].width=25

    ws1.column_dimensions['A'].width=20
    ws1.column_dimensions['B'].width=20
    ws1.column_dimensions['C'].width=20
    ws1.column_dimensions['D'].width=20
    ws1.column_dimensions['E'].width=20
    ws1.column_dimensions['F'].width=20
    ws1.column_dimensions['G'].width=20
    ws1.column_dimensions['H'].width=20
    ws1.column_dimensions['I'].width=20
    ws1.column_dimensions['J'].width=20
    ws1.column_dimensions['K'].width=20
    ws1.column_dimensions['L'].width=20
    ws1.column_dimensions['M'].width=20
    ws1.column_dimensions['N'].width=20
    ws1.column_dimensions['O'].width=20
    ws1.column_dimensions['P'].width=20
    ws1.column_dimensions['Q'].width=20
    ws1.column_dimensions['R'].width=20
    ws1.column_dimensions['S'].width=20
    ws1.column_dimensions['T'].width=20
    ws1.column_dimensions['U'].width=20
    ws1.column_dimensions['V'].width=20
    ws1.column_dimensions['W'].width=20

    ws7.column_dimensions['A'].width=20
    ws7.column_dimensions['B'].width=20
    ws7.column_dimensions['C'].width=20
    ws7.column_dimensions['D'].width=20
    ws7.column_dimensions['E'].width=20
    ws7.column_dimensions['F'].width=20
    ws7.column_dimensions['G'].width=20
    ws7.column_dimensions['H'].width=20
    ws7.column_dimensions['I'].width=20
    ws7.column_dimensions['J'].width=20
    ws7.column_dimensions['K'].width=20
    ws7.column_dimensions['L'].width=20


    ws10.column_dimensions['A'].width=50
    ws10.column_dimensions['B'].width=20
    ws10.column_dimensions['C'].width=25
    ws10.column_dimensions['D'].width=20

    

    border_a=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    position=['A4','B4','C4','D4','E4','F4','G4','H4',
              'A5','B5','C5','D5','E5','F5','G5','H5',
              'E7','F7','G7','H7','E8','F8','G8','H8',
              'A12','B12','C12','D12','E12','F12','G12','H12',
              'A29','B29','C29','D29',
              'A75','B75','C75']
    for cell in position:
        ws6[cell].border=border_a


    border_b=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    position= ['A13','B13','C13','D13','E13','F13','G13','H13',
       	      'A14','B14','C14','D14','E14','F14','G14','H14',
       	      'A15','B15','C15','D15','E15','F15','G15','H15',
       	      'A16','B16','C16','D16','E16','F16','G16','H16',
              'A17','B17','C17','D17','E17','F17','G17','H17',
              'A18','B18','C18','D18','E18','F18','G18','H18',
              'A19','B19','C19','D19','E19','F19','G19','H19',
              'A20','B20','C20','D20','E20','F20','G20','H20',
              'A21','B21','C21','D21','E21','F21','G21','H21',
              'A22','B22','C22','D22','E22','F22','G22','H22',
              'A23','B23','C23','D23','E23','F23','G23','H23',
              'A24','B24','C24','D24','E24','F24','G24','H24',
              'A25','B25','C25','D25','E25','F25','G25','H25',
              'A26','B26','C26','D26','E26','F26','G26','H26',
              'A30','B30','C30','D30',
              'A31','B31','C31','D31',
              'A32','B32','C32','D32',
              'A33','B33','C33','D33',
              'A34','B34','C34','D34',
              'A35','B35','C35','D35',
              'A36','B36','C36','D36',
              'A37','B37','C37','D37',
              'A38','B38','C38','D38',
              'A39','B39','C39','D39',
              'A40','B40','C40','D40',
              'A41','B41','C41','D41',
              'A42','B42','C42','D42',
              'A43','B43','C43','D43',
              'A47','B47','C47','D47','E47','F47',
              'A48','B48','C48','D48','E48','F48',
              'A49','B49','C49','D49','E49','F49',
              'A50','B50','C50','D50','E50','F50',
              'A51','B51','C51','D51','E51','F51',
              'A52','B52','C52','D52','E52','F52',
              'A53','B53','C53','D53','E53','F53',
              'A54','B54','C54','D54','E54','F54',
              'A55','B55','C55','D55','E55','F55',
              'A56','B56','C56','D56','E56','F56',
              'A57','B57','C57','D57','E57','F57',
              'A58','B58','C58','D58','E58','F58',
              'A59','B59','C59','D59','E59','F59',
              'A60','B60','C60','D60','E60','F60',
              'A61','B61','C61','D61','E61','F61',
              'A62','B62','C62','D62','E62','F62',
              'A63','B63','C63','D63','E63','F63',
              'A64','B64','C64','D64','E64','F64',
              'A65','B65','C65','D65','E65','F65',
              'A66','B66','C66','D66','E66','F66',
              'A67','B67','C67','D67','E67','F67',
              'A68','B68','C68','D68','E68','F68',
              'A69','B69','C69','D69','E69','F69',
              'A70','B70','C70','D70','E70','F70',
              'A71','B71','C71','D71','E71','F71',
              'A72','B72','C72','D72','E72','F72',
              'A76','B76','C76',
              'A77','B77','C77',
              'A78','B78','C78',
              'A79','B79','C79',
              'A80','B80','C80',
              'A81','B81','C81',
              'A82','B82','C82',
              'A83','B83','C83',
              'A84','B84','C84']





    for	cell in	position:
        ws6[cell].border=border_b 


    ws6['A44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))



    ws6['F33'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F34'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G33'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G34'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))


    font_bold=Font(bold=True)
    position= ['A4','B4','C4','D4','E4','F4','G4','H4',
                'E7','F7','G7','H7',
                'A1','C1','A3','A9','A11',
                'F33','F34']
    for cell in position:
        ws6[cell].font=font_bold



    ws6['A1'].font=Font(size=16)
    ws6['C1'].font=Font(size=16)
    ws6['H1'].font=Font(size=16)

    ws6['A1'].border=Border(left=Side(border_style=BORDER_THICK), right=Side(border_style=BORDER_THICK), top=Side(border_style=BORDER_THICK), bottom=Side(border_style=BORDER_THICK))
    ws6['C1'].border=Border(left=Side(border_style=BORDER_THICK), right=Side(border_style=BORDER_THICK), top=Side(border_style=BORDER_THICK), bottom=Side(border_style=BORDER_THICK))


    colour= PatternFill("solid", fgColor="00CCFFFF")
    position= ['A4','B4','C4','D4','E4','F4', 'G4','H4','E7','F7','G7','H7', 'F33','F34']
    for cell in position:
        ws6[cell].fill=colour


    colour= PatternFill("solid", fgColor="009999FF")
    position= ['A12','B12','C12','D12','E12','F12','G12','H12']
    for cell in position:
        ws6[cell].fill=colour
    

    colour=PatternFill("solid", fgColor="DCDCDC")
    position=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1']
    for cell in position:
        ws1[cell].fill=colour



    border_c=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    position=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1']
    for cell in position:
        ws1[cell].border=border_c


    font_bold==Font(bold=True)
    position=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1']
    for cell in position:
        ws1[cell].font=font_bold


    ws2['U9']= "Y/N"

    colour=PatternFill("solid", fgColor="DCDCDC")
    position=['A9','B9','C9','D9','E9','F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9','R9','S9','T9','U9']
    for cell in position:
        ws2[cell].fill=colour


    font_bold=Font(bold=True)
    position=['A9','B9','C9','D9','E9','F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9','R9','S9','T9','U9',
              'B3','B6','E3','E6','G3','G6','K3','K6','I3']
    for cell in position:
        ws2[cell].font=font_bold


    border_d=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    position= ['B9','C9','D9','E9','F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9','Q9','R9','S9','T9']
    for cell in position:
        ws2[cell].border=border_d


    ws2['A9'].border=Border(left=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['U9'].border=Border(right=Side(border_style=BORDER_MEDIUM),top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    font_bold=Font(bold=True)
    position=['B2','C2','D2','E2','F2','G2','H2','I2','J2','K2','L2','M2','N2']
    for cell in position:
        ws4[cell].font=font_bold


    font_bold=Font(bold=True)
    position=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1']
    for	cell in	position:
        ws7[cell].font=font_bold


    wb.save(path+sampleid+'_'+referral+'_panCancer_both.xlsx')



if __name__ == "__main__":

    
    #Insert information
    runid=sys.argv[1]
    sampleid=sys.argv[2]
    worksheet=sys.argv[3]
    referral=sys.argv[4]

    print(runid)
    print(sampleid)
    print(worksheet)
    print(referral)
    

    path="/data/results/"+runid + "/RochePanCancer/"


    referral=referral.upper()
    if referral=="BREAST":
        referral="Breast"
    elif referral=="COLORECTAL":
        referral="Colorectal"
    elif referral== "DPYD":
        referral="DPYD"
    elif referral=="GIST":
        referral="GIST"
    elif referral=="GLIOMA":
        referral="Glioma"
    elif referral=="HEADANDNECK":
        referral="HeadAndNeck"
    elif referral=="LUNG":
        referral="Lung"
    elif referral=="MELANOMA":
        referral="Melanoma"
    elif referral=="OVARIAN":
        referral="Ovarian"
    elif referral=="PROSTATE":
        referral="Prostate"
    elif referral=="THYROID":
        referral="Thyroid"
    elif referral=="TUMOUR":
        referral="Tumour"
    else:
        print ("referral not recognised")    
    

    referrals_list=['Breast','Colorectal','DPYD','GIST','Glioma','HeadAndNeck','Lung','Melanoma','Ovarian','Prostate','Thyroid', 'Tumour']

    referral_present=False
    
    for referral_value in referrals_list:
        if (referral==referral_value):
            referral_present=True
 
    num_rows_coverage=0

    if (referral_present==True):
    
        variant_report_NTC=get_variantReport_NTC(referral, path)
 
        variant_report_referral=get_variant_report(referral, path, sampleid)

        variant_report_NTC_2=add_extra_columns_NTC_report(variant_report_NTC, variant_report_referral)

        variant_report_referral_2=expand_variant_report(variant_report_referral, variant_report_NTC_2)
        
        if (referral!="GIST" and referral!="DPYD"):
            CNV_file=get_CNV_file(referral, path, sampleid)

        

        coverage_value="250x"

        gaps_file=get_gaps_file(referral, path, sampleid, coverage_value)

        hotspots_coverage=get_hotspots_coverage_file(referral, path, sampleid, coverage_value)

        hotspots_coverage_NTC=get_NTC_hotspots_coverage_file(referral, path, coverage_value)

        hotspots_coverage_2, num_rows_coverage=add_columns_hotspots_coverage(hotspots_coverage, hotspots_coverage_NTC)
   

        if (referral!="GIST" and referral != "DPYD"):

            genescreen_coverage=get_genescreen_coverage_file(referral, path, sampleid, coverage_value)
     
            genescreen_coverage_NTC= get_NTC_genescreen_coverage_file(referral, path, coverage_value)

            genescreen_coverage_2=add_columns_genescreen_coverage(genescreen_coverage, genescreen_coverage_NTC, num_rows_coverage)

        subpanel_coverage=get_subpanel_coverage(referral, path, sampleid, coverage_value)



        coverage_value="135x"
        gaps_file=get_gaps_file(referral, path, sampleid, coverage_value)


        variant_report_referral_3=match_polys_and_artefacts(variant_report_referral_2, variant_report_NTC_2)

        add_excel_formulae()

    else:
        print("referral not in referrals_list")
