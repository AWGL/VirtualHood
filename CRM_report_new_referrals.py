
from openpyxl import Workbook
import pandas
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import xlrd
import sys
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THIN, BORDER_THICK
from openpyxl.styles import Font
import argparse

wb=Workbook()
ws1= wb.create_sheet("Sheet_1")
ws9= wb.create_sheet("Sheet_9")
ws10= wb.create_sheet("Sheet_10")
ws2= wb.create_sheet("Sheet_2")
ws4= wb.create_sheet("Sheet_4")
ws5= wb.create_sheet("Sheet_5")
ws6= wb.create_sheet("Sheet_6")
ws7= wb.create_sheet("Sheet_7")

#name the tabs
ws1.title="Patient demographics"
ws2.title="Variant_calls"
ws4.title="Mutations and SNPs"
ws5.title="hotspots.gaps"
ws6.title="Report"
ws7.title="NTC variant"
ws9.title="Subpanel NTC check"
ws10.title="Subpanel coverage"


#set the page layout of the report
ws6.page_setup.orientation=ws6.ORIENTATION_LANDSCAPE
ws6.page_setup.paperSize=ws6.PAPERSIZE_A4

#Add titles to patient demographics tab
ws1['A1']='Date Received'
ws1['B1']='LAB No'
ws1['C1']='Name'
ws1['D1']='Tumour %'
ws1['E1']='Analysis'
ws1['F1']='Panel run'
ws1['G1']='Qubit [DNA] ng/ul'
ws1['H1']='Dilution (ng/ul)'
ws1['I1']='NGS wks'
ws1['J1']='Date Set up'
ws1['K1']='Date of MiSeq run'
ws1['L1']='Library ng/ul'
ws1['M1']='Library nM'

#variant calls table headers
ws2['B3']='DNA number'
ws2['B6']='Patient name'

ws2['E3']='NTC check 1'
ws2['E6']='NTC check 2'

ws2['G3']='1st checker name & date'
ws2['G6']='2nd checker name & date'

ws2['K3']='CRM worksheet'
ws2['K6']='Analysis pipeline:CRM'
ws2['A8']=" "

ws2['I3']= "% Tumour"


#Mutations and SNPs table headers
ws4['B2']='Gene'
ws4['C2']='Exon/Intron'
ws4['D2']='HGVS c'
ws4['E2']='HGVS p'
ws4['F2']='Allele Frequency'
ws4['G2']='Quality'
ws4['H2']='Depth'
ws4['I2']='Classification'
ws4['J2']='Transcript'
ws4['K2']='Variant'
ws4['L2']='Position'
ws4['M2']='Conclusion 1st checker'
ws4['N2']='Conclusion 2nd checker'


#Report headers
ws6['A3']='Patient information'
ws6['A4']='Lab number'
ws6['B4']='Patient Name'
ws6['C4']='Tumour %'
ws6['D4']='Analysis'
ws6['E4']='Qubit[DNA] ng/ul'
ws6['F4']='Dilution (ng/ul)'


ws6['E7']='NGS wks'
ws6['F7']='Date set up'
ws6['G7']='Date of MiSeq run'
ws6['H7']='Library ng/ul(Qubit)'
ws6['I7']='Library nm'
ws6['J7']='NTC check 1'
ws6['K7']='NTC check 2'
ws6['E7']='NGS wks'
ws6['F7']='Average molarity(nM)'
ws6['G7']='NextSeq run ID'

ws6['A4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['B4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['C4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['D4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['E4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['F4'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['E7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['F7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['G7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['H7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['I7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['J7'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['K7'].fill= PatternFill("solid", fgColor="00CCFFFF")


ws6['H46'].fill= PatternFill("solid", fgColor="00CCFFFF")
ws6['H47'].fill= PatternFill("solid", fgColor="00CCFFFF")

ws6['A11']='Confirmed variant calls'
ws6['A12']='Gene'
ws6['B12']='Exon'
ws6['C12']='Variant'
ws6['D12']='HGVS c.'
ws6['E12']='HGVS p.'
ws6['F12']='Allele frequency'
ws6['G12']='Conclusion 1st checker'
ws6['H12']='Conclusion 2nd checker'

ws6['A12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['B12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['C12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['D12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['E12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['F12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['G12'].fill= PatternFill("solid", fgColor="009999FF")
ws6['H12'].fill= PatternFill("solid", fgColor="009999FF")

    


def get_variantReport_NTC(referral, path, NTC_name, runid): 
    '''
    Fill out the NTC variants tab using the relevant variant report
    '''

    if(os.stat(path+NTC_name+"/hotspot_variants/"+runid+"_"+NTC_name+"_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report_NTC=pandas.read_csv(path+NTC_name+"/hotspot_variants/"+runid+"_"+NTC_name+"_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(path+NTC_name+"/hotspot_variants/"+runid+"_"+NTC_name+"_"+referral+"_VariantReport.txt")
    else:
        variant_report_NTC=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "Preferred","dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
 

    variant_report_NTC_2=pandas.DataFrame(variant_report_NTC)

    #Sort by preferred transcript and filter columns
    variant_report_NTC_3=variant_report_NTC_2[variant_report_NTC_2.Preferred!=False]
    variant_report_NTC_4= variant_report_NTC_3.iloc[:,[23,29,25,26,2,5,3,6,24,1]]

    return (variant_report_NTC_4)




def get_variant_report(referral, path, sampleid, runid):

    '''
    Open the relevant variant file to append to the variant calls tab.
    '''

    if(os.stat(path+ sampleid +"/hotspot_variants/"+runid+"_"+sampleid+"_"+referral+"_VariantReport.txt").st_size!=0):
        variant_report=pandas.read_csv(path+sampleid+"/hotspot_variants/"+runid+"_" +sampleid+"_"+referral+"_VariantReport.txt", sep="\t")
        ws6['A9']=(sampleid+"_"+referral+"_VariantReport.txt")
    else:
        variant_report=pandas.DataFrame(columns=["SampleID", "Variant", "Filter", "Frequency", "Depth", "Genotype", "Quality", "Classification", "Preferred", "dbSNP", "Cosmic", "HGMD", "ExAC_African","ExAC_American", "ExAC_EuropeanNonFinnish", "ExAC_Finnish", "ExAC_EastAsian", "ExAC_SouthAsian", "ExAC_Other", "1KG_African", "1KG_American","1KG_European", "1KG_EastAsian", "1KG_SouthAsian", "Gene", "Transcript", "HGVSc", "HGVSp", "Consequence", "INTRON", "EXON", "SIFT", "PolyPhen"])
            

    #sort dataframe by preferred transcript an filter out certain columns
    variant_report_2=pandas.DataFrame(variant_report)
    variant_report_3=variant_report_2[variant_report_2.Preferred!=False]
    variant_report_4= variant_report_3.iloc[:,[23,29,25,26,2,5,3,6,24,1]]    
    
    #Add position column in variant calls tab by splitting the variant column 
    variant_list=[]    
    for variant in variant_report_4['Variant']:
        variant_2=[]
        value=0
        for char in variant:
            if(value==0):
                if (char.isdigit()==True or char=="X" or char==":"):
                    variant_2.append(char)
                else:
                    value=1
        variant_3=''.join(variant_2)
        variant_list.append(variant_3)
        

    variant_report_4['Position']=variant_list

    
    return (variant_report_4)





def add_extra_columns_NTC_report(variant_report_NTC_4, variant_report_4, ws7,wb, path,):
    
    '''
    Add 'Present in sample' and 'variant allele calls' columns to NTC variant table
    '''

    num_rows_NTC=variant_report_NTC_4.shape[0]
    num_rows_variant_report=variant_report_4.shape[0]

    variant_in_sample=[]
    row=0
    while (row<num_rows_NTC):
        row2=0
        present='NO'
        while (row2<num_rows_variant_report):
            if (variant_report_4.iloc[row2,9]==variant_report_NTC_4.iloc[row,9]):
                present='YES'
            row2=row2+1
        variant_in_sample.append(present)
        row=row+1
     
   
    variant_report_NTC_4['Present in sample']=variant_in_sample


    variant_allele_calls=[]
    row=0
    num_rows_NTC=variant_report_NTC_4.shape[0]
    while (row<num_rows_NTC):
        variant_report_NTC_4.iloc[row,6]=int(variant_report_NTC_4.iloc[row,6])
        allele_call= variant_report_NTC_4.iloc[row,4]*variant_report_NTC_4.iloc[row,6]
        variant_allele_calls.append(allele_call)
        row=row+1

    variant_report_NTC_4['Variant allele calls']=variant_allele_calls    

    for row in dataframe_to_rows(variant_report_NTC_4, header=True, index=False):
        ws7.append(row)


    return (variant_report_NTC_4, ws7)




def expand_variant_report(variant_report_4, variant_report_NTC_4):

    '''
    create the extra table at the side of the variant_calls tab
    '''

    detection_threshold=[]
    variant_in_NTC=[]


    row=0
    num_rows_variant_report=variant_report_4.shape[0]


    while (row<num_rows_variant_report):
        variant_report_4.iloc[row,6]= int(variant_report_4.iloc[row,6])
        if (variant_report_4.iloc[row,6]<=500):
            value_2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            value_2=str(value_2)
        else:
            value_2=0
        detection_threshold.append(value_2)
        row2=0
        num_rows_NTC=variant_report_NTC_4.shape[0]
        present='NO'
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row,9]==variant_report_NTC_4.iloc[row2,9]):
                present='YES'
            row2=row2+1
        variant_in_NTC.append(present)
        row=row+1


    variant_report_4["Conclusion 1st checker"]=""
    variant_report_4["QC"]=""
    variant_report_4["Conclusion 2nd checker"]=""
    variant_report_4["QC "]=""


    variant_report_4["Detection threshold based on depth"]=detection_threshold
    variant_report_4["Is variant present in NTC "]=variant_in_NTC

    return (variant_report_4)




def get_gaps_file(referral, path, sampleid, ws5, wb, runid):
 
    '''
    Open the relevant gap file to append to the end of the mutations and snps tab. If the gap file is empty, write 'no gaps'.
    '''

    if(os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_" +sampleid+"_"+referral+".gaps").st_size==0):
        ws5['A1']= 'No gaps'
        gaps=""
    if (os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_" +sampleid+"_"+referral+".gaps").st_size!=0):
        gaps=pandas.read_csv(path+ sampleid+"/hotspot_coverage/"+runid+"_" +sampleid+"_"+referral+".gaps", sep="\t")
        for row in dataframe_to_rows(gaps, header=True, index=False):
            ws5.append(row)


    ws6['C30']="=hotspots.gaps!D1"
    ws6['C31']="=hotspots.gaps!D2"
    ws6['C32']="=hotspots.gaps!D3"
    ws6['C33']="=hotspots.gaps!D4"
    ws6['C34']="=hotspots.gaps!D5"
    ws6['C35']="=hotspots.gaps!D6"
    ws6['C36']="=hotspots.gaps!D7"
    ws6['C37']="=hotspots.gaps!D8"
    ws6['C38']="=hotspots.gaps!D9"
    ws6['C39']="=hotspots.gaps!D10"
    ws6['C40']="=hotspots.gaps!D11"
    ws6['C41']="=hotspots.gaps!D12"
    ws6['C42']="=hotspots.gaps!D13"
    ws6['C43']="=hotspots.gaps!D14"
    ws6['C44']="=hotspots.gaps!D15"


    ws6['D30']="=hotspots.gaps!D16"
    ws6['D31']="=hotspots.gaps!D17"
    ws6['D32']="=hotspots.gaps!D18"
    ws6['D33']="=hotspots.gaps!D19"
    ws6['D34']="=hotspots.gaps!D20"
    ws6['D35']="=hotspots.gaps!D21"
    ws6['D36']="=hotspots.gaps!D22"
    ws6['D37']="=hotspots.gaps!D23"
    ws6['D38']="=hotspots.gaps!D24"
    ws6['D39']="=hotspots.gaps!D25"
    ws6['D40']="=hotspots.gaps!D26"
    ws6['D41']="=hotspots.gaps!D27"
    ws6['D42']="=hotspots.gaps!D28"
    ws6['D43']="=hotspots.gaps!D29"
    ws6['D44']="=hotspots.gaps!D30"

    ws6['E30']="=hotspots.gaps!D31"
    ws6['E31']="=hotspots.gaps!D32"
    ws6['E32']="=hotspots.gaps!D33"
    ws6['E33']="=hotspots.gaps!D34"
    ws6['E34']="=hotspots.gaps!D35"
    ws6['E35']="=hotspots.gaps!D36"
    ws6['E36']="=hotspots.gaps!D37"
    ws6['E37']="=hotspots.gaps!D38"
    ws6['E38']="=hotspots.gaps!D39"
    ws6['E39']="=hotspots.gaps!D40"
    ws6['E40']="=hotspots.gaps!D41"
    ws6['E41']="=hotspots.gaps!D42"
    ws6['E42']="=hotspots.gaps!D43"
    ws6['E43']="=hotspots.gaps!D44"
    ws6['E44']="=hotspots.gaps!D45"

    ws6['F30']="=hotspots.gaps!D46"
    ws6['F31']="=hotspots.gaps!D47"
    ws6['F32']="=hotspots.gaps!D48"
    ws6['F33']="=hotspots.gaps!D49"
    ws6['F34']="=hotspots.gaps!D50"
    ws6['F35']="=hotspots.gaps!D51"
    ws6['F36']="=hotspots.gaps!D52"
    ws6['F37']="=hotspots.gaps!D53"
    ws6['F38']="=hotspots.gaps!D54"
    ws6['F39']="=hotspots.gaps!D55"
    ws6['F40']="=hotspots.gaps!D56"
    ws6['F41']="=hotspots.gaps!D57"
    ws6['F42']="=hotspots.gaps!D58"
    ws6['F43']="=hotspots.gaps!D59"
    ws6['F44']="=hotspots.gaps!D60"

    ws6['G30']="=hotspots.gaps!D61"
    ws6['G31']="=hotspots.gaps!D62"
    ws6['G32']="=hotspots.gaps!D63"
    ws6['G33']="=hotspots.gaps!D64"
    ws6['G34']="=hotspots.gaps!D65"
    ws6['G35']="=hotspots.gaps!D66"
    ws6['G36']="=hotspots.gaps!D67"
    ws6['G37']="=hotspots.gaps!D68"
    ws6['G38']="=hotspots.gaps!D69"
    ws6['G39']="=hotspots.gaps!D70"
    ws6['G40']="=hotspots.gaps!D71"
    ws6['G41']="=hotspots.gaps!D72"
    ws6['G42']="=hotspots.gaps!D73"
    ws6['G43']="=hotspots.gaps!D74"
    ws6['G44']="=hotspots.gaps!D75"

    ws6['H30']="=hotspots.gaps!D76"
    ws6['H31']="=hotspots.gaps!D77"
    ws6['H32']="=hotspots.gaps!D78"
    ws6['H33']="=hotspots.gaps!D79"
    ws6['H34']="=hotspots.gaps!D80"
    ws6['H35']="=hotspots.gaps!D81"
    ws6['H36']="=hotspots.gaps!D82"
    ws6['H37']="=hotspots.gaps!D83"
    ws6['H38']="=hotspots.gaps!D84"
    ws6['H39']="=hotspots.gaps!D85"
    ws6['H40']="=hotspots.gaps!D86"
    ws6['H41']="=hotspots.gaps!D87"
    ws6['H42']="=hotspots.gaps!D88"
    ws6['H43']="=hotspots.gaps!D89"
    ws6['H44']="=hotspots.gaps!D90"

    ws6['H46']="Analysed by:"
    ws6['H47']="Checked by:"


    return (gaps, ws5)




def get_hotspots_coverage_file(referral, path, sampleid, runid):

    '''
    Open the relevant coverage file to append to the end of the mutations and snps tab. If the coverage file is empty, write 'No hotspots'.
    '''

    if(os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_"+referral+".coverage").st_size==0):
        ws9['A1']= 'No hotspots'
    if (os.stat(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_"+referral+".coverage").st_size!=0):
        Coverage=pandas.read_csv(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_"+referral+".coverage", sep="\t")
      

    Coverage= Coverage.iloc[:,[3,4,5]]

    return(Coverage)




def get_NTC_hotspots_coverage_file(referral, path, NTC_name, runid):

    '''
    Open the relevant NTC hotspots coverage file.
    '''
       
    if(os.stat(path+ NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_"+referral+".coverage").st_size==0):
        data= [{'CHR':'NA', 'START':'NA', 'END':'NA', 'META':'NA', 'AVG_DEPTH':'NA', 'PERC_COVERAGE@250':'NA'}]
        NTC_check=pandas.DataFrame(data)
    if (os.stat(path+ NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_"+referral+".coverage").st_size!=0):
        NTC_check=pandas.read_csv(path+ NTC_name+"/hotspot_coverage/"+runid+"_"+NTC_name+"_"+referral+".coverage", sep="\t")
    
    return(NTC_check)




def add_columns_hotspots_coverage(Coverage, NTC_check, ws9):

    #Add percentage NTC and subpanel columns to the Coverage table
    
    Coverage['NTC_AVG_Depth']=""
    Coverage['%NTC']=""

    num_rows_NTC= NTC_check.shape[0]
    num_rows_sample= Coverage.shape[0]

    row1=0

    while(row1<num_rows_sample):
        row2=0
        while (row2<num_rows_NTC):
            if(Coverage.iloc[row1,0] == NTC_check.iloc[row2,3]):
                Coverage.iloc[row1,3]= NTC_check.iloc[row2,4]
            row2=row2+1
        row1=row1+1
    

    Coverage['%NTC']=Coverage['NTC_AVG_Depth']/Coverage['AVG_DEPTH']
    Coverage['%NTC']= Coverage['%NTC']*100

    for row in dataframe_to_rows(Coverage, header=True, index=False):
        ws9.append(row)

    num_rows_coverage=Coverage.shape[0]
    row =0
    while (row< num_rows_coverage):
        row_spreadsheet=row+2
        row_spreadsheet_2=str(row_spreadsheet)
        if (Coverage.iloc[row,4]>10):
            ws9['E'+row_spreadsheet_2].fill= PatternFill("solid", fgColor="FFBB00")
        row=row+1
    
    return (Coverage, num_rows_coverage,ws9)




def get_subpanel_coverage(referral, path, sampleid, runid, ws10):

    #Add coverage table
    if(os.stat(path+sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_coverage.txt").st_size==0):
        ws10['A1']= 'No coverage'
        Coverage=""
        Coverage_2=""
    if (os.stat(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_coverage.txt").st_size!=0):
        Coverage=pandas.read_csv(path+ sampleid+"/hotspot_coverage/"+runid+"_"+sampleid+"_coverage.txt", sep="\t")

        s=Coverage['FEATURE'].apply(lambda x: x.split('_'))
        Coverage['Referral']=s.apply(lambda x:x[5])
        Coverage_2=Coverage[Coverage.Referral==referral]
      
        for row in dataframe_to_rows(Coverage_2, header=True, index=False):
            ws10.append(row)
    

    return(Coverage_2, ws10)




def match_polys_and_artefacts(variant_report_4, variant_report_NTC_4, artefacts_path, ws2):

    '''
    Extract the relevant information from "PanCancer_Poly and Artefact list.xlsx" by matching the variant name with the ones in the variant report table
    '''

    poly_artefact_dict={}
    poly_and_Artefact_list=pandas.read_excel(artefacts_path +"CRM_poly_artefact_list.xlsx")
    poly_and_Artefact_list_2=pandas.DataFrame(poly_and_Artefact_list)


    num_rows_variant_report=variant_report_4.shape[0]
    num_rows_poly_artefact=poly_and_Artefact_list_2.shape[0]


    #Fill the conclusion columns using the relevant column in the Poly and Artefact spreadsheet
    row1=0
    while (row1<num_rows_variant_report):
        row2=0
        while(row2<num_rows_poly_artefact):
            if (poly_and_Artefact_list_2.iloc[row2,0]==variant_report_4.iloc[row1,9]):
                variant_report_4.iloc[row1,11]= poly_and_Artefact_list_2.iloc[row2,9]
                variant_report_4.iloc[row1,13]= poly_and_Artefact_list_2.iloc[row2,9]
            row2=row2+1
        row1=row1+1


    #fill second table of variant-calls tab using the conclusion column of the first table
    row3=0
    while (row3<num_rows_variant_report):
        if (variant_report_4.iloc[row3,11]=='Known artefact'):
            variant_report_4.iloc[row3,12]=3
            variant_report_4.iloc[row3,14]=3
        if (variant_report_4.iloc[row3,11]=='Known Poly'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc[row3,14]=1
        if (variant_report_4.iloc[row3,11]=='WT'):
            variant_report_4.iloc[row3,12]=3
            variant_report_4.iloc[row3,14]=3
        if (variant_report_4.iloc[row3,11]=='Genuine'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc_[row3,14]=1
        if (variant_report_4.iloc[row3,11]=='SNP'):
            variant_report_4.iloc[row3,12]=1
            variant_report_4.iloc_[row3,14]=1

        row3=row3+1

 

   #Add extra columns to the variant report table to determine level of NTC contamination

    
    variant_report_4["#of mutant reads in patient sample "]=""
    variant_report_4["#of mutant reads in NTC if present "]=""
    variant_report_4["Is the NTC contamination significant?"]=""
    variant_report_4["Y/N"]=""



    num_rows_NTC= variant_report_NTC_4.shape[0]
    row=0


    while (row<num_rows_variant_report):
        if variant_report_4.iloc[row,16]=="YES":
            variant_report_4.iloc[row,6]=float(variant_report_4.iloc[row,6])
            variant_report_4.iloc[row,4]=float(variant_report_4.iloc[row,4])
            value2= variant_report_4.iloc[row,4]*variant_report_4.iloc[row,6]
            variant_report_4.iloc[row,17]=value2
        row2=0
        while (row2<num_rows_NTC):
            if (variant_report_4.iloc[row, 9]==variant_report_NTC_4.iloc[row2,9]):
                variant_report_4.iloc[row,18]=variant_report_NTC_4.iloc[row2,11]
                variant_report_4.iloc[row,19]=variant_report_4.iloc[row,18]/variant_report_4.iloc[row,17]
            row2=row2+1
        row=row+1



    #Add upper-limit and lower-limit variant report dataframes to the excel workbook
    
    variant_report_4_upper_limit=variant_report_4[variant_report_4.Frequency>0.045]


    #Add how conclusion was reached column

    variant_report_4_upper_limit['Comments/Notes/evidence:how conclusion was reached']=""

    row=0

    num_rows_variant_report_upper_limit=variant_report_4_upper_limit.shape[0]

    while (row<num_rows_variant_report_upper_limit):
        if (variant_report_4_upper_limit.iloc[row,11]=='Known artefact'):
            variant_report_4_upper_limit.iloc[row,21]='On artefact list'
        if (variant_report_4_upper_limit.iloc[row,11]=='Known Poly'):
            variant_report_4_upper_limit.iloc[row,21]='On Poly list'
        if (variant_report_4_upper_limit.iloc[row,11]=='WT'):
            variant_report_4_upper_limit.iloc[row,21]='SNP in Ref.Seq'
        row=row+1



    for row in dataframe_to_rows(variant_report_4_upper_limit, header=True, index=False):
        ws2.append(row)



    return(variant_report_4, ws2)



def add_excel_formulae(wb, ws1, ws2, ws4, ws5, ws6, ws7, ws9, ws10):

    #add excel formulae to the spreadsheets to enable automation after program has finished

    ws2['I4']= "='Patient demographics'!D2"
     
    ws6['A13']= "='Mutations and SNPS'!B3"
    ws6['B13']= "='Mutations and SNPS'!C3"
    ws6['C13']= "='Mutations and SNPS'!K3"
    ws6['D13']= "='Mutations and SNPS'!D3"
    ws6['E13']= "='Mutations and SNPS'!E3"
    ws6['F13']= "='Mutations and SNPS'!F3"
    ws6['G13']= "='Mutations and SNPS'!M3"
    ws6['H13']= "='Mutations and SNPS'!N3"

    ws6['A14']= "='Mutations and SNPS'!B4"
    ws6['B14']= "='Mutations and SNPS'!C4"
    ws6['C14']= "='Mutations and SNPS'!K4"
    ws6['D14']= "='Mutations and SNPS'!D4"
    ws6['E14']= "='Mutations and SNPS'!E4"
    ws6['F14']= "='Mutations and SNPS'!F4"
    ws6['G14']= "='Mutations and SNPS'!M4"
    ws6['H14']= "='Mutations and SNPS'!N4"

    ws6['A15']= "='Mutations and SNPS'!B5"
    ws6['B15']= "='Mutations and SNPS'!C5"
    ws6['C15']= "='Mutations and SNPS'!K5"
    ws6['D15']= "='Mutations and SNPS'!D5"
    ws6['E15']= "='Mutations and SNPS'!E5"
    ws6['F15']= "='Mutations and SNPS'!F5"
    ws6['G15']= "='Mutations and SNPS'!M5"
    ws6['H15']= "='Mutations and SNPS'!N5"

    ws6['A16']= "='Mutations and SNPS'!B6"
    ws6['B16']= "='Mutations and SNPS'!C6"
    ws6['C16']= "='Mutations and SNPS'!K6"
    ws6['D16']= "='Mutations and SNPS'!D6"
    ws6['E16']= "='Mutations and SNPS'!E6"
    ws6['F16']= "='Mutations and SNPS'!F6"
    ws6['G16']= "='Mutations and SNPS'!M6"
    ws6['H16']= "='Mutations and SNPS'!N6"

    ws6['A17']= "='Mutations and SNPS'!B7"
    ws6['B17']= "='Mutations and SNPS'!C7"
    ws6['C17']= "='Mutations and SNPS'!K7"
    ws6['D17']= "='Mutations and SNPS'!D7"
    ws6['E17']= "='Mutations and SNPS'!E7"
    ws6['F17']= "='Mutations and SNPS'!F7"
    ws6['G17']= "='Mutations and SNPS'!M7"
    ws6['H17']= "='Mutations and SNPS'!N7"

    ws6['A18']= "='Mutations and SNPS'!B8"
    ws6['B18']= "='Mutations and SNPS'!C8"
    ws6['C18']= "='Mutations and SNPS'!K8"
    ws6['D18']= "='Mutations and SNPS'!D8"
    ws6['E18']= "='Mutations and SNPS'!E8"
    ws6['F18']= "='Mutations and SNPS'!F8"
    ws6['G18']= "='Mutations and SNPS'!M8"
    ws6['H18']= "='Mutations and SNPS'!N8"

    ws6['A19']= "='Mutations and SNPS'!B9"
    ws6['B19']= "='Mutations and SNPS'!C9"
    ws6['C19']= "='Mutations and SNPS'!K9"
    ws6['D19']= "='Mutations and SNPS'!D9"
    ws6['E19']= "='Mutations and SNPS'!E9"
    ws6['F19']= "='Mutations and SNPS'!F9"
    ws6['G19']= "='Mutations and SNPS'!M9"
    ws6['H19']= "='Mutations and SNPS'!N9"

    ws6['A20']= "='Mutations and SNPS'!B10"
    ws6['B20']= "='Mutations and SNPS'!C10"
    ws6['C20']= "='Mutations and SNPS'!K10"
    ws6['D20']= "='Mutations and SNPS'!D10"
    ws6['E20']= "='Mutations and SNPS'!E10"
    ws6['F20']= "='Mutations and SNPS'!F10"
    ws6['G20']= "='Mutations and SNPS'!M10"
    ws6['H20']= "='Mutations and SNPS'!N10"

    ws6['A21']= "='Mutations and SNPS'!B11"
    ws6['B21']= "='Mutations and SNPS'!C11"
    ws6['C21']= "='Mutations and SNPS'!K11"
    ws6['D21']= "='Mutations and SNPS'!D11"
    ws6['E21']= "='Mutations and SNPS'!E11"
    ws6['F21']= "='Mutations and SNPS'!F11"
    ws6['G21']= "='Mutations and SNPS'!M11"
    ws6['H21']= "='Mutations and SNPS'!N11"

    ws6['A22']= "='Mutations and SNPS'!B12"
    ws6['B22']= "='Mutations and SNPS'!C12"
    ws6['C22']= "='Mutations and SNPS'!K12"
    ws6['D22']= "='Mutations and SNPS'!D12"
    ws6['E22']= "='Mutations and SNPS'!E12"
    ws6['F22']= "='Mutations and SNPS'!F12"
    ws6['G22']= "='Mutations and SNPS'!M12"
    ws6['H22']= "='Mutations and SNPS'!N12"

    ws6['A23']= "='Mutations and SNPS'!B13"
    ws6['B23']= "='Mutations and SNPS'!C13"
    ws6['C23']= "='Mutations and SNPS'!K13"
    ws6['D23']= "='Mutations and SNPS'!D13"
    ws6['E23']= "='Mutations and SNPS'!E13"
    ws6['F23']= "='Mutations and SNPS'!F13"
    ws6['G23']= "='Mutations and SNPS'!M13"
    ws6['H23']= "='Mutations and SNPS'!N13"

    ws6['A24']= "='Mutations and SNPS'!B14"
    ws6['B24']= "='Mutations and SNPS'!C14"
    ws6['C24']= "='Mutations and SNPS'!K14"
    ws6['D24']= "='Mutations and SNPS'!D14"
    ws6['E24']= "='Mutations and SNPS'!E14"
    ws6['F24']= "='Mutations and SNPS'!F14"
    ws6['G24']= "='Mutations and SNPS'!M14"
    ws6['H24']= "='Mutations and SNPS'!N14"

    ws6['A25']= "='Mutations and SNPS'!B15"
    ws6['B25']= "='Mutations and SNPS'!C15"
    ws6['C25']= "='Mutations and SNPS'!K15"
    ws6['D25']= "='Mutations and SNPS'!D15"
    ws6['E25']= "='Mutations and SNPS'!E15"
    ws6['F25']= "='Mutations and SNPS'!F15"
    ws6['G25']= "='Mutations and SNPS'!M15"
    ws6['H25']= "='Mutations and SNPS'!N15"

    ws6['A26']= "='Mutations and SNPS'!B16"
    ws6['B26']= "='Mutations and SNPS'!C16"
    ws6['C26']= "='Mutations and SNPS'!K16"
    ws6['D26']= "='Mutations and SNPS'!D16"
    ws6['E26']= "='Mutations and SNPS'!E16"
    ws6['F26']= "='Mutations and SNPS'!F16"
    ws6['G26']= "='Mutations and SNPS'!M16"
    ws6['H26']= "='Mutations and SNPS'!N16"


    ws6['A5'] = sampleid
    ws6['B5']="='Patient demographics'!C2"
    ws6['C5']="='Patient demographics'!D2"
    ws6['D5']= referral
    ws6['E5']= "='Patient demographics'!G2"
    ws6['F5']= "='Patient demographics'!H2"
    ws6['E8']="='Patient demographics'!I2"
    ws6['F8']="='Patient demographics'!J2"
    ws6['G8']="='Patient demographics'!K2"
    ws6['H8']="='Patient demographics'!L2"
    ws6['I8']="='Patient demographics'!M2"
    ws6['J8']="='Variant_calls'!E4"
    ws6['K8']="='Variant_calls'!E7"


    ws6['A30']="='Subpanel coverage'!A2"
    ws6['A31']="='Subpanel coverage'!A3"
    ws6['A32']= "='Subpanel coverage'!A4"
    ws6['A33']="='Subpanel coverage'!A5"
    ws6['A34']="='Subpanel coverage'!A6"
    ws6['A35']="='Subpanel coverage'!A7"
    ws6['A36']="='Subpanel coverage'!A8"
    ws6['A37']="='Subpanel coverage'!A9"
    ws6['A38']="='Subpanel coverage'!A10"
    ws6['A39']="='Subpanel coverage'!A11"
    ws6['A40']="='Subpanel coverage'!A12"
    ws6['A41']="='Subpanel coverage'!A13"
    ws6['A42']="='Subpanel coverage'!A14"
    ws6['A43']="='Subpanel coverage'!A15"

    ws6['B30']="='Subpanel coverage'!C2"
    ws6['B31']="='Subpanel coverage'!C3"
    ws6['B32']="='Subpanel coverage'!C4"
    ws6['B33']="='Subpanel coverage'!C5"
    ws6['B34']="='Subpanel coverage'!C6"
    ws6['B35']="='Subpanel coverage'!C7"
    ws6['B36']="='Subpanel coverage'!C8"
    ws6['B37']="='Subpanel coverage'!C9"
    ws6['B38']="='Subpanel coverage'!C10"
    ws6['B39']="='Subpanel coverage'!C11"
    ws6['B40']= "='Subpanel coverage'!C12"
    ws6['B41']="='Subpanel coverage'!C13"
    ws6['B42']="='Subpanel coverage'!C14"
    ws6['B43']="='Subpanel coverage'!C15"


    ws6['A29']= sampleid +"_" + referral
    ws6['C29'] = "Gaps in hotspots ROI"

    ws6['A27']=sampleid
    ws6['A27'].font= Font(bold=True)
    ws6['A27'].font=Font(size=16)

    ws6['A29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['B29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['C29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['D29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['E29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['F29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['G29'].fill= PatternFill("solid", fgColor="FFBB00")
    ws6['H29'].fill= PatternFill("solid", fgColor="FFBB00")

    ws6['A1']=sampleid
    ws6['C1']='Patient Analysis Summary Sheet-CRM'

        

    ws9['J4']="NTC check 1"
    ws9['J5']="NTC check 2"

    ws2['B4']= sampleid
    ws2['B7']="='Patient demographics'!C2"
    ws2['E4']="='Subpanel NTC check'!K4"
    ws2['E7']="='Subpanel NTC check'!K5"
    ws2['K4']=worksheet


    #change widths of columns
    ws6.column_dimensions['C'].width=40
    ws6.column_dimensions['A'].width=40
    ws6.column_dimensions['B'].width=15
    ws6.column_dimensions['D'].width=10
    ws6.column_dimensions['E'].width=20
    ws6.column_dimensions['F'].width=20
    ws6.column_dimensions['G'].width=20
    ws6.column_dimensions['H'].width=20
    ws9.column_dimensions['A'].width=60
    ws9.column_dimensions['B'].width=15
    ws9.column_dimensions['C'].width=25
    ws9.column_dimensions['D'].width=20
    ws9.column_dimensions['E'].width=15
    ws9.column_dimensions['F'].width=15
    ws2.row_dimensions[9].height=40
    ws2.row_dimensions[61].height=40

    ws2.column_dimensions['C'].width=20
    ws2.column_dimensions['D'].width=53
    ws2.column_dimensions['E'].width=20
    ws2.column_dimensions['H'].width=20
    ws2.column_dimensions['I'].width=20
    ws2.column_dimensions['J'].width=20
    ws2.column_dimensions['K'].width=20
    ws2.column_dimensions['L'].width=20
    ws2.column_dimensions['M'].width=20
    ws2.column_dimensions['N'].width=20
    ws2.column_dimensions['O'].width=20
    ws2.column_dimensions['P'].width=40
    ws2.column_dimensions['Q'].width=30
    ws2.column_dimensions['R'].width=33
    ws2.column_dimensions['S'].width=33
    ws2.column_dimensions['T'].width=40
    ws2.column_dimensions['V'].width=50


    ws4.column_dimensions['B'].width=20
    ws4.column_dimensions['C'].width=20
    ws4.column_dimensions['D'].width=20
    ws4.column_dimensions['E'].width=20
    ws4.column_dimensions['F'].width=20
    ws4.column_dimensions['G'].width=20
    ws4.column_dimensions['H'].width=20
    ws4.column_dimensions['I'].width=20
    ws4.column_dimensions['J'].width=20
    ws4.column_dimensions['K'].width=20
    ws4.column_dimensions['L'].width=20
    ws4.column_dimensions['M'].width=25
    ws4.column_dimensions['N'].width=25

    ws1.column_dimensions['A'].width=20
    ws1.column_dimensions['B'].width=20
    ws1.column_dimensions['C'].width=20
    ws1.column_dimensions['D'].width=20
    ws1.column_dimensions['E'].width=20
    ws1.column_dimensions['F'].width=20
    ws1.column_dimensions['G'].width=20
    ws1.column_dimensions['H'].width=20
    ws1.column_dimensions['I'].width=20
    ws1.column_dimensions['J'].width=20
    ws1.column_dimensions['K'].width=20
    ws1.column_dimensions['L'].width=20
    ws1.column_dimensions['M'].width=20

    ws7.column_dimensions['A'].width=20
    ws7.column_dimensions['B'].width=20
    ws7.column_dimensions['C'].width=20
    ws7.column_dimensions['D'].width=20
    ws7.column_dimensions['E'].width=20
    ws7.column_dimensions['F'].width=20
    ws7.column_dimensions['G'].width=20
    ws7.column_dimensions['H'].width=20
    ws7.column_dimensions['I'].width=20
    ws7.column_dimensions['J'].width=20
    ws7.column_dimensions['K'].width=20
    ws7.column_dimensions['L'].width=20


    ws10.column_dimensions['A'].width=80
    ws10.column_dimensions['B'].width=20
    ws10.column_dimensions['C'].width=25
    ws10.column_dimensions['D'].width=20


    ws6['A4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F4'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['A5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F5'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['E7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['J7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['K7'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['J8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['K8'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))  

    ws6['A12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H12'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))


    ws6['A29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H29'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['A13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H13'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H14'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H15'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H16'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))


    ws6['A17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H17'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H18'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H19'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H20'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H21'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H22'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H23'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H24'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H25'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
 
    ws6['A26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H26'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H30'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
   
    ws6['A31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G32'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H31'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H33'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H34'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H35'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H36'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H37'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H38'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H39'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN)) 

    ws6['A40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H40'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H41'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H42'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['B43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['C43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['D43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['E43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['F43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['G43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws6['H43'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws6['A44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['B44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['C44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['D44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['E44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['F44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['G44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H44'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['H46'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['H47'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I46'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws6['I47'].border=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws6['A4'].font= Font(bold=True)
    ws6['B4'].font= Font(bold=True)
    ws6['C4'].font= Font(bold=True)
    ws6['D4'].font= Font(bold=True)
    ws6['E4'].font= Font(bold=True)
    ws6['F4'].font= Font(bold=True)
    ws6['G4'].font= Font(bold=True)
    ws6['H4'].font= Font(bold=True)
    ws6['I4'].font= Font(bold=True)

    ws6['E7'].font= Font(bold=True)
    ws6['F7'].font= Font(bold=True)
    ws6['G7'].font= Font(bold=True)
    ws6['H7'].font= Font(bold=True)
    ws6['I7'].font= Font(bold=True)
    ws6['J7'].font= Font(bold=True)
    ws6['K7'].font= Font(bold=True)

    ws6['A1'].font= Font(bold=True)
    ws6['C1'].font= Font(bold=True)
    ws6['A3'].font= Font(bold=True)
    ws6['A9'].font= Font(bold=True)
    ws6['A11'].font= Font(bold=True)

    ws6['H46'].font= Font(bold=True)
    ws6['H47'].font= Font(bold=True)

    ws6['A1'].font=Font(size=16)
    ws6['C1'].font=Font(size=16)
    ws6['H1'].font=Font(size=16)

    ws6['A1'].border=Border(left=Side(border_style=BORDER_THICK), right=Side(border_style=BORDER_THICK), top=Side(border_style=BORDER_THICK), bottom=Side(border_style=BORDER_THICK))
    ws6['C1'].border=Border(left=Side(border_style=BORDER_THICK), right=Side(border_style=BORDER_THICK), top=Side(border_style=BORDER_THICK), bottom=Side(border_style=BORDER_THICK))

    ws1['A1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['B1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['C1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['D1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['E1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['F1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['G1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['H1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['I1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['J1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['K1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['L1'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws1['M1'].fill= PatternFill("solid", fgColor="DCDCDC")


    ws1['A1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['B1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['C1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['D1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['E1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['F1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['G1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['H1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['I1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['J1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['K1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['L1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))
    ws1['M1'].border=Border(left=Side(border_style=BORDER_THIN), right=Side(border_style=BORDER_THIN), top=Side(border_style=BORDER_THIN), bottom=Side(border_style=BORDER_THIN))

    ws1['A1'].font= Font(bold=True)
    ws1['B1'].font= Font(bold=True)
    ws1['C1'].font= Font(bold=True)
    ws1['D1'].font= Font(bold=True)
    ws1['E1'].font= Font(bold=True)
    ws1['F1'].font= Font(bold=True)
    ws1['G1'].font= Font(bold=True)
    ws1['H1'].font= Font(bold=True)
    ws1['I1'].font= Font(bold=True)
    ws1['J1'].font= Font(bold=True)
    ws1['K1'].font= Font(bold=True)
    ws1['L1'].font= Font(bold=True)
    ws1['M1'].font= Font(bold=True)
 
    ws2['A9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['B9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['C9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['D9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['E9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['F9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['G9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['H9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['I9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['J9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['K9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['L9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['M9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['N9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['O9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['P9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['Q9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['R9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['S9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['T9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['U9'].fill= PatternFill("solid", fgColor="DCDCDC")
    ws2['V9'].fill= PatternFill("solid", fgColor="DCDCDC")

    ws2['A9'].font= Font(bold=True)
    ws2['B9'].font= Font(bold=True)
    ws2['C9'].font= Font(bold=True)
    ws2['D9'].font= Font(bold=True)
    ws2['E9'].font= Font(bold=True)
    ws2['F9'].font= Font(bold=True)
    ws2['G9'].font= Font(bold=True)
    ws2['H9'].font= Font(bold=True)
    ws2['I9'].font= Font(bold=True)
    ws2['J9'].font= Font(bold=True)
    ws2['K9'].font= Font(bold=True)
    ws2['L9'].font= Font(bold=True)
    ws2['M9'].font= Font(bold=True)
    ws2['N9'].font= Font(bold=True)
    ws2['O9'].font= Font(bold=True)
    ws2['P9'].font= Font(bold=True)
    ws2['Q9'].font= Font(bold=True)
    ws2['R9'].font= Font(bold=True)
    ws2['S9'].font= Font(bold=True)
    ws2['T9'].font= Font(bold=True)
    ws2['U9'].font= Font(bold=True)
    ws2['V9'].font= Font(bold=True)


    ws2['B3'].font= Font(bold=True)
    ws2['B6'].font= Font(bold=True)
    ws2['E3'].font= Font(bold=True)
    ws2['E6'].font= Font(bold=True)
    ws2['G3'].font= Font(bold=True)
    ws2['G6'].font= Font(bold=True)
    ws2['K3'].font= Font(bold=True)
    ws2['K6'].font= Font(bold=True)
    ws2['I3'].font= Font(bold=True)

    ws2['A9'].border=Border(left=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['B9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['C9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['D9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['E9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['F9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['G9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['H9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))    
    ws2['I9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['J9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['K9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['L9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['M9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['N9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['O9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['P9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['Q9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['R9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['S9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['T9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['U9'].border=Border(top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
    ws2['V9'].border=Border(right=Side(border_style=BORDER_MEDIUM),top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))

    ws4['B2'].font= Font(bold=True)
    ws4['C2'].font= Font(bold=True)
    ws4['D2'].font= Font(bold=True)
    ws4['E2'].font= Font(bold=True)
    ws4['F2'].font= Font(bold=True)
    ws4['G2'].font= Font(bold=True)
    ws4['H2'].font= Font(bold=True)
    ws4['I2'].font= Font(bold=True)
    ws4['J2'].font= Font(bold=True)
    ws4['K2'].font= Font(bold=True)
    ws4['L2'].font= Font(bold=True)
    ws4['M2'].font= Font(bold=True)
    ws4['N2'].font= Font(bold=True)

    ws7['A1'].font= Font(bold=True)
    ws7['B1'].font= Font(bold=True)
    ws7['C1'].font= Font(bold=True)
    ws7['D1'].font= Font(bold=True)
    ws7['E1'].font= Font(bold=True)
    ws7['F1'].font= Font(bold=True)
    ws7['G1'].font= Font(bold=True)
    ws7['H1'].font= Font(bold=True)
    ws7['I1'].font= Font(bold=True)
    ws7['J1'].font= Font(bold=True)
    ws7['K1'].font= Font(bold=True)
    ws7['L1'].font= Font(bold=True)

    wb.save(path+sampleid+'_'+referral+'_CRM.xlsx')



if __name__ == "__main__":

    
    #Insert information
    parser=argparse.ArgumentParser()
    parser.add_argument('--runid', required=True)
    parser.add_argument('--sampleid', required=True)
    parser.add_argument('--worksheet', required=True)
    parser.add_argument('--referral', required=True)
    parser.add_argument('--NTC_name', required=True)
    parser.add_argument('--path', required=False)
    parser.add_argument('--artefacts', required=False)
    args=parser.parse_args()

    runid=args.runid
    sampleid=args.sampleid
    worksheet=args.worksheet
    referral=args.referral
    NTC_name=args.NTC_name
    path=args.path
    artefacts_path=args.artefacts

    if (path==None):
        path="/data/results/"+runid + "/NGHS-101X/"
    if (artefacts_path==None):
        artefacts_path="/data/temp/artefacts_lists/"


    referral=referral.upper()
    if referral=="COLORECTAL":
        referral="Colorectal"
    elif referral=="GIST":
        referral="GIST"
    elif referral=="GLIOMA":
        referral="Glioma"
    elif referral=="LUNG":
        referral="Lung"
    elif referral=="MELANOMA":
        referral="Melanoma"
    elif referral == "THYROID":
        referral = "Thyroid"
    elif referral=="TUMOUR":
        referral="Tumour"
    else:
        print ("referral not recognised")    
    

    referrals_list=['Colorectal', 'GIST', 'Glioma', 'Lung', 'Melanoma', 'Thyroid', 'Tumour']

    referral_present=False
    
    for referral_value in referrals_list:
        if (referral==referral_value):
            referral_present=True
 
    num_rows_coverage=0

    if (referral_present==True):
    
        variant_report_NTC=get_variantReport_NTC(referral, path, NTC_name, runid)
 
        variant_report_referral=get_variant_report(referral, path, sampleid, runid)

        variant_report_NTC_2, ws7=add_extra_columns_NTC_report(variant_report_NTC, variant_report_referral, ws7, wb, path)

        variant_report_referral_2=expand_variant_report(variant_report_referral, variant_report_NTC_2)

        gaps_file, ws5=get_gaps_file(referral, path, sampleid, ws5, wb, runid)

        hotspots_coverage=get_hotspots_coverage_file(referral, path, sampleid, runid)

        hotspots_coverage_NTC=get_NTC_hotspots_coverage_file(referral, path, NTC_name, runid)

        hotspots_coverage_2, num_rows_coverage, ws9=add_columns_hotspots_coverage(hotspots_coverage, hotspots_coverage_NTC, ws9)
   
        subpanel_coverage, ws10=get_subpanel_coverage(referral, path, sampleid, runid, ws10)

        variant_report_referral_3, ws2=match_polys_and_artefacts(variant_report_referral_2, variant_report_NTC_2, artefacts_path, ws2)

        add_excel_formulae(wb, ws1, ws2, ws4, ws5, ws6, ws7, ws9, ws10)

    else:
        print("referral not in referrals_list")
